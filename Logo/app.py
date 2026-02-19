# -*- coding: utf-8 -*-
import json
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime

import pandas as pd
import streamlit as st
from pandas.io.formats.style import Styler
import io
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

st.markdown(
    """
    <link rel="apple-touch-icon" sizes="180x180" href="Logo/apple-touch-icon.png">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="default">
    <meta name="apple-mobile-web-app-title" content="Lefties vs Righties">
    """,
    unsafe_allow_html=True,
)



APP_NAME = "Lefties vs Righties Ryder Cup"
APP_VERSION = "1.2.12"
APP_CREATED = "19.02.2026"

DATA_FILE = "Data/GolfData.xlsx"
FAVICON_FILE = "Logo/LR.ico"
PLAYERS_PATH = "Players/"
ANONYM_FILE = PLAYERS_PATH + "Anonym.jpg"  # fallback portrét, keď hráč nemá fotku alebo URL neexistuje
BASE_DIR = Path(__file__).resolve().parent
PLAYERS_DIR = BASE_DIR / 'Players'


# -----------------------------
# Detekcia zariadenia / OS (User-Agent) + rozlíšenie (JS)
# -----------------------------

def detect_device_os() -> tuple[str, str, str]:
    """Vráti (device, os_name, raw_user_agent).

    User-Agent čítame z `st.context.headers` (ak je dostupné). Ak nie, vrátime 'Neznáme'.
    """
    ua = ""
    try:
        ctx = getattr(st, 'context', None)
        if ctx is not None and getattr(ctx, 'headers', None) is not None:
            ua = ctx.headers.get('User-Agent', '') or ''
    except Exception:
        ua = ""

    ua_l = ua.lower() if isinstance(ua, str) else ""

    device = "Neznáme zariadenie"
    if "iphone" in ua_l:
        device = "iPhone"
    elif "ipad" in ua_l:
        device = "iPad"
    elif "android" in ua_l and "mobile" in ua_l:
        device = "Android telefón"
    elif "android" in ua_l:
        device = "Android tablet"
    elif "windows" in ua_l:
        device = "PC (Windows)"
    elif "macintosh" in ua_l or "mac os x" in ua_l:
        device = "Mac"
    elif "linux" in ua_l:
        device = "Linux"

    import re
    os_name = "Neznámy OS"
    if "iphone" in ua_l or "ipad" in ua_l or "ipod" in ua_l:
        m = re.search(r"os (\d+[_\.]\d+(?:[_\.]\d+)?)", ua_l)
        ver = m.group(1).replace('_', '.') if m else ""
        os_name = f"iOS {ver}".strip() if ver else "iOS"
    elif "android" in ua_l:
        m = re.search(r"android (\d+(?:\.\d+)*)", ua_l)
        ver = m.group(1) if m else ""
        os_name = f"Android {ver}".strip() if ver else "Android"
    elif "windows nt" in ua_l:
        m = re.search(r"windows nt (\d+(?:\.\d+)*)", ua_l)
        ver = m.group(1) if m else ""
        os_name = f"Windows NT {ver}".strip() if ver else "Windows"
    elif "mac os x" in ua_l:
        m = re.search(r"mac os x (\d+[\._]\d+(?:[\._]\d+)?)", ua_l)
        ver = m.group(1).replace('_', '.') if m else ""
        os_name = f"macOS {ver}".strip() if ver else "macOS"
    elif "linux" in ua_l:
        os_name = "Linux"

    return device, os_name, ua


def get_display_metrics() -> dict:
    """Získa viewport/screen/dpr z prehliadača.

    Používa balík `streamlit-javascript` (import `streamlit_javascript`).
    Ak nie je dostupný alebo blokovaný (napr. embed/iframe policy), vráti {}.
    """
    try:
        from streamlit_javascript import st_javascript  # pip install streamlit-javascript
    except Exception:
        return {}

    try:
        import json
        # vraciame JSON string; niektoré verzie vracajú priamo dict
        payload = st_javascript(
            "JSON.stringify({iw: window.innerWidth, ih: window.innerHeight, dpr: window.devicePixelRatio, sw: screen.width, sh: screen.height})",
            key="__display_metrics__",
        )
        if not payload:
            return {}
        data = json.loads(payload) if isinstance(payload, str) else payload

        def _to_int(v):
            try:
                return int(round(float(v)))
            except Exception:
                return None

        def _to_float(v):
            try:
                return float(v)
            except Exception:
                return None

        out = {
            'inner_width': _to_int(data.get('iw')),
            'inner_height': _to_int(data.get('ih')),
            'screen_width': _to_int(data.get('sw')),
            'screen_height': _to_int(data.get('sh')),
            'dpr': _to_float(data.get('dpr')),
        }
        if out.get('inner_width') and out.get('inner_height') and out.get('dpr'):
            out['physical_viewport_width'] = int(round(out['inner_width'] * out['dpr']))
            out['physical_viewport_height'] = int(round(out['inner_height'] * out['dpr']))
        return out
    except Exception:
        return {}


def classify_device_type(device_label: str, metrics: dict) -> str:
    """Určí typ zariadenia (mobil/tablet/desktop).

    Primárne podľa viewport šírky (CSS px). Fallback podľa UA detekcie.
    """
    w = metrics.get('inner_width') if isinstance(metrics, dict) else None
    if isinstance(w, int) and w > 0:
        if w <= 600:
            return 'mobil'
        if w <= 1024:
            return 'tablet'
        return 'desktop'

    d = (device_label or '').lower()
    if 'iphone' in d or 'android telefón' in d:
        return 'mobil'
    if 'ipad' in d or 'android tablet' in d:
        return 'tablet'
    return 'desktop'


st.set_page_config(
    page_title=APP_NAME, layout="wide",
    page_icon=FAVICON_FILE
)

# --- UI: odstránenie prázdneho priestoru nad hlavičkou (logo čo najvyššie) ---
st.markdown(
    """
    <style>
    /* TOP PADDING RESET */
    [data-testid="stAppViewContainer"] .main .block-container { padding-top: 0rem !important; }
    [data-testid="stAppViewContainer"] .main { padding-top: 0rem !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <style>
    /* ===== APP HEADER (MOBILE + DESKTOP) ===== */
    .app-header-mobile{width:100%; text-align:center; margin: 0.25rem 0 0.35rem;}
    .app-title-mobile{width:100%; font-weight:900; line-height:1.05; font-size:1.75rem; display:block;}
    .app-version-mobile{width:100%; color:#666; font-size:0.92rem; margin-top:0.15rem;}
    .app-header-desktop{display:flex; align-items:center; gap:16px; margin:8px 0 6px;}
    .app-logo-desktop{height:64px; width:auto; display:block;}
    .app-title-desktop{font-size:1.75rem; font-weight:800; line-height:1.05; margin:0 0 2px 0;}
    .app-version-desktop{color:#666; font-size:0.95rem; line-height:1.0; margin:0;}
    </style>
    """,
    unsafe_allow_html=True,
)


# --- UI: mobile tabuľky nech sa zmestia na šírku (menší font) ---
st.markdown(
    """
    <style>
    /* MOBILE TABLE FIT */
    .mobile-fit table { width: 100% !important; table-layout: fixed !important; }
    .mobile-fit th, .mobile-fit td { padding: 0.20rem 0.25rem !important; }
    .mobile-fit table { font-size: 0.78rem !important; }
    .mobile-fit td { word-wrap: break-word; overflow-wrap: anywhere; }
    </style>
    """,
    unsafe_allow_html=True,
)


# -- Vlastné štýly INLINE (eliminácia styles.css)  ### REPLACE
STYLES_INLINE = r"""
/* ===== SORT LIŠTA – kľúč: zasiahnúť parent "stHorizontalBlock" ===== */
.stats-sortbar [data-testid="stHorizontalBlock"] {
  /* celé pole stĺpcov budeme brať ako jeden kompaktný riadok */
  display: inline-flex !important;
  flex-wrap: nowrap !important;
  align-items: center !important;
  justify-content: flex-start !important;
  gap: 0 !important;                 /* ZRUŠ medzery medzi stĺpcami */
  white-space: nowrap;
}

/* každý stĺpec má mať šírku len podľa obsahu */
.stats-sortbar [data-testid="column"] {
  flex: 0 0 auto !important;         /* žiadne percentá, žiadne rozťahovanie */
  width: auto !important;
  padding: 0 !important;
  margin: 0 !important;
  min-width: 0 !important;
}

/* wrapper stButton bez okrajov */
.stats-sortbar div[data-testid="stButton"] {
  margin: 0 !important;
}

/* samotné tlačidlo = kompaktné, bez roztiahnutia */
.stats-sortbar div[data-testid="stButton"] > button {
  width: auto !important;
  display: inline-block;
  white-space: nowrap;
  padding: 0.35rem 0.75rem;
  border-radius: 6px;
  margin: 0 !important;
}

/* farbenie podľa markeru (ponechané) */
.marker { display:block; height:0; margin:0; padding:0; }
.marker.win-left + div[data-testid="stButton"] > button {
  background-color:#cfe8ff !important; border:1px solid #80808033; color:#111 !important;
}
.marker.win-right + div[data-testid="stButton"] > button {
  background-color:#ffd6d6 !important; border:1px solid #80808033; color:#111 !important;
}
.marker.win-left  + div[data-testid="stButton"] > button:hover { background-color:#b7dcff !important; }
.marker.win-right + div[data-testid="stButton"] > button:hover { background-color:#ffc2c2 !important; }

/* pevná medzera medzi skupinami – cez tvoj token `sep` */
.stats-sortbar .sep {
  display: inline-block;
  width: 0.75rem;  /* uprav podľa potreby */
  height: 1px;
}

/* (voliteľné) ak je viewport veľmi úzky, povoľ horizontálny scroll */
@media (max-width:560px){
  .stats-sortbar [data-testid="stHorizontalBlock"] { overflow-x: auto; }
}
"""
st.markdown(f"<style>{STYLES_INLINE}</style>", unsafe_allow_html=True)

# -- Dodatočné štýly: aktívne triediace tlačidlo = tučné + väčšie písmo
st.markdown(
    """
<style>
/***** marker pred tlačidlom *****/
.marker { display:block; height:0; margin:0; padding:0; }
/***** Aktívne triediace tlačidlo (robustné selektory) *****/
.marker.sort-active + div[data-testid="stButton"] > button,
.marker.sort-active + div[data-testid="stButton"] button,
.marker.sort-active + div [data-testid="baseButton-secondary"],
.marker.sort-active + div button {
  font-weight: 700 !important;
  font-size: 1.05rem !important; /* zladené s tabuľkou */
}
</style>
""",
    unsafe_allow_html=True,
)

# -- Sticky hlavička (2 riadky) + scroll kontajner 600px
st.markdown(
    """
<style>
/* Kontajner so scrollom pre tabuľku štatistík */
.sticky-table-container {
  max-height: 600px;     /* požadovaná výška viewportu pre tabuľku */
  overflow: auto;        /* zvyšok scrolluje */
}

/* Stabilné lepenie hlavičky */
.sticky-table-container table {
  border-collapse: separate;  /* dôležité pre position: sticky */
  border-spacing: 0;
}

/* 1. riadok hlavičky (level 0) */
.sticky-table-container thead th.col_heading.level0 {
  position: sticky;
  top: 0;
  z-index: 3;
  background: #fff;
}
/* orientačná výška 1. riadku hlavičky */
.sticky-table-container thead tr:nth-child(1) th.col_heading.level0 {
  height: 36px;
}

/* 2. riadok hlavičky (level 1) */
.sticky-table-container thead th.col_heading.level1 {
  position: sticky;
  top: 36px;   /* zhodné s výškou 1. riadku */
  z-index: 4;
  background: #fff;
}

/* Bold + centrovanie */
.sticky-table-container thead th {
  font-weight: 700 !important;
  text-align: center !important;
}
</style>
""",
    unsafe_allow_html=True,
)

SORT_SELECT_CSS = r"""
/* Kontajner držíme kompaktný – len tak široký, ako je obsah */
.sort-select {
  display: inline-block;            /* kľúčové: neťahá sa na 100% riadku */
  max-width: 100%;                  /* bezpečnostná poistka */
}

/* Streamlit selectbox (parent wrapper) */
.sort-select [data-testid="stSelectbox"] {
  width: fit-content !important;    /* nech je podľa obsahu */
  min-width: 0 !important;
}

/* Vnútorný BaseWeb <Select> wrapper */
.sort-select [data-baseweb="select"] {
  width: fit-content !important;
  min-width: 0 !important;
}

/* Niektoré buildy majú ešte jeden medziobal <div> – poistka */
.sort-select [data-testid="stSelectbox"] > div {
  width: fit-content !important;
}

/* (voliteľné) label – ak ho zobrazuješ, nech je tesnejšie nad selectom */
.sort-select [data-testid="stSelectbox"] label {
  margin-bottom: 0.25rem;
}
"""
st.markdown(f"<style>{SORT_SELECT_CSS}</style>", unsafe_allow_html=True)

STATS_FIT_CSS = r"""
/* Kontajner pre selectbox + tabuľku: nech je tak široký, ako jeho obsah (tabuľka) */
.stats-fit {
  display: inline-block;      /* kľúčové: šírka = najširšie dieťa (tabuľka) */
  max-width: 100%;            /* poistka, aby to nikdy nepretieklo mimo layout */
  vertical-align: top;        /* estetika pri susedných blokoch */
}

/* Tabuľka nech používa svoju prirodzenú šírku */
.stats-fit table {
  width: auto !important;     /* nesťahuj na 100% kontajnera */
}

/* Selectbox nech sa natiahne na 100% šírky .stats-fit (teda = šírka tabuľky) */
.stats-fit [data-testid="stSelectbox"],
.stats-fit [data-testid="stSelectbox"] > div,
.stats-fit [data-baseweb="select"],
.stats-fit [data-baseweb="select"] > div {
  width: 100% !important;
  min-width: 0 !important;
}

/* (voliteľné) trocha miesta pod selectom */
.stats-fit [data-testid="stSelectbox"] {
  margin-bottom: 0.5rem;
}
"""
st.markdown(f"<style>{STATS_FIT_CSS}</style>", unsafe_allow_html=True)

# -- Farby tímov
COLOR_LEFT_BG = "#E6F2FF"  # bledomodrá
COLOR_RIGHT_BG = "#FCE8E8"  # bledočervená
# -- URL loga
RAW_LOGO_URL = "https://raw.githubusercontent.com/Jasen77/lefties-righties/main/Logo/logo.png"

@st.cache_data(show_spinner=False)
def load_data(xlsx_path: str):
    xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
    df_matches = pd.read_excel(xls, sheet_name="Zápasy", engine="openpyxl")
    df_tournaments = pd.read_excel(xls, sheet_name="Turnaje", engine="openpyxl")
    return df_matches, df_tournaments

@st.cache_data(show_spinner=False)
def load_players_sheet(xlsx_path: str) -> pd.DataFrame:
    """
    Načíta hárok 'Hráči' s menami a portrétmi.
    Ošetrí aj variant názvu stĺpca 'Portrét'/'Portret' a z buniek vyextrahuje prvú http(s) URL.
    """
    try:
        xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
        if "Hráči" not in xls.sheet_names:
            return pd.DataFrame()
        dfp = pd.read_excel(xls, sheet_name="Hráči", engine="openpyxl")

        # Normalize názvy stĺpcov (niekde býva 'Portret', inde 'Portrét')
        cols = {c: str(c).strip() for c in dfp.columns}
        dfp.rename(columns=cols, inplace=True)

        # nájdi portrétový stĺpec
        portrait_col = None
        for cand in ("Portrét", "Portret"):
            if cand in dfp.columns:
                portrait_col = cand
                break
        if portrait_col is None or "Hráč" not in dfp.columns:
            return pd.DataFrame()  # chýbajú kľúčové stĺpce

        # vytiahni prvú http(s) URL z bunky (ak je tam hypertext/poznámka)
        def _first_url(v) -> str | None:
            if pd.isna(v):
                return None
            s = str(v)
            m = re.search(r"https?://\S+", s)
            return m.group(0).strip(")];,") if m else None

        dfp["_portrait_raw"] = dfp[portrait_col]
        dfp["_portrait_url"] = dfp[portrait_col].apply(_first_url)
        # doplň aj relatívne cesty (Players/...) na raw github URL
        dfp["_portrait_url"] = dfp["_portrait_url"].fillna(dfp["_portrait_raw"].apply(lambda v: (str(v).strip() if pd.notna(v) else None)))
        # _portrait_url ostáva len ako prípadná URL vyextrahovaná z bunky
        dfp["_portrait_url"] = dfp["_portrait_url"]
        # kľúč 'Hráč' nechávame v kanonickom formáte, zhoduje sa s menami v L1/L2/R1/R2
        return dfp[["Hráč", "_portrait_url", "_portrait_raw"]].copy()
    except Exception:
        return pd.DataFrame()
        

if not Path(DATA_FILE).exists():
    st.error(f"Nebolo možné nájsť súbor {DATA_FILE} v aktuálnom adresári.")
    st.stop()

# -- DÁTA
df_matches, df_tournaments = load_data(DATA_FILE)
df_players_sheet = load_players_sheet(DATA_FILE)

# --- Detekcia prostredia (pre layout hlavičky) ---
_device, _os_name, _ua = detect_device_os()
_metrics = get_display_metrics()
_device_type = classify_device_type(_device, _metrics)

# --- Header: logo + názov + verzia (kompaktnejšie medzery) ---
if _device_type == "mobil":
    # mobil: logo na celú šírku obrazovky, pod ním názov a verzia (názov cez celú šírku)
    st.image(RAW_LOGO_URL, use_container_width=True)
    st.markdown(
        f"""
        <div class="app-header-mobile">
          <span class="app-title-mobile">{APP_NAME}</span>
          <span class="app-version-mobile">ver.: {APP_VERSION} ({APP_CREATED})</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
else:
    # desktop/tablet: logo + názov + verzia v jednom riadku
    st.markdown(
        f"""
        <div class="app-header-desktop">
          <img src="{RAW_LOGO_URL}" alt="Logo aplikácie" class="app-logo-desktop"/>
          <div>
            <div class="app-title-desktop">{APP_NAME}</div>
            <div class="app-version-desktop">ver.: {APP_VERSION} ({APP_CREATED})</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# --- Prostredie: zariadenie, OS a rozlíšenie (zobrazenie) ---
# _device, _os_name, _ua = detect_device_os()
# _metrics = get_display_metrics()
# _device_type = classify_device_type(_device, _metrics)

_res_parts = []
if _metrics.get('inner_width') and _metrics.get('inner_height'):
    _res_parts.append(f"Viewport: **{_metrics['inner_width']}×{_metrics['inner_height']}** CSS px")
if _metrics.get('dpr'):
    _res_parts.append(f"DPR: **{_metrics['dpr']:.2g}**")
if _metrics.get('physical_viewport_width') and _metrics.get('physical_viewport_height'):
    _res_parts.append(f"Odhad px (viewport×DPR): **{_metrics['physical_viewport_width']}×{_metrics['physical_viewport_height']}** px")

_res_text = " | ".join(_res_parts) if _res_parts else "Rozlíšenie: **nedostupné** (skontroluj `streamlit-javascript`)"

# Na stránke (viditeľné okamžite)
# st.info(f"Zariadenie: **{_device}** ({_device_type}) | OS: **{_os_name}** | {_res_text}")
  # zapo
# V sidebar-e (aby to bolo vždy viditeľné)
with st.sidebar:
    st.markdown("### Prostredie")
    st.write(f"Zariadenie: {_device} ({_device_type})")
    st.write(f"OS: {_os_name}")
    if _metrics:
        st.write(f"Viewport: {_metrics.get('inner_width')}×{_metrics.get('inner_height')} CSS px")
        st.write(f"DPR: {_metrics.get('dpr')}")
    else:
        st.caption("Rozlíšenie nie je dostupné (JS komponent).")

    # --- Obnoviť dáta: vymaže cache a reštartuje appku
    if st.button('🔄 Obnoviť dáta', use_container_width=True, key='btn_refresh_data'):
        try:
            st.cache_data.clear()
        except Exception:
            pass
        try:
            st.cache_resource.clear()
        except Exception:
            pass
        try:
            st.session_state.clear()
        except Exception:
            pass
        st.rerun()



# with st.expander("Technické detaily (User-Agent / displej)"):
#     st.code(_ua or "User-Agent nebol dostupný (Streamlit neposkytol hlavičky).")
#     if _metrics:
#         st.json(_metrics)
#     else:
#         st.warning("Displejové metriky nie sú dostupné. Tip: spusti appku cez `python -m streamlit run app.py` a over `pip show streamlit-javascript` v tom istom prostredí.")
# 
# -----------------------------
# Helpers
# -----------------------------

def _clean_name(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    return s if s else None


def to_firstname_first(name: str) -> str:
    """Z 'Priezvisko Meno' urobí 'Meno Priezvisko'."""
    if not isinstance(name, str):
        return name
    parts = name.split()
    if len(parts) < 2:
        return name
    first = parts[-1]
    last = " ".join(parts[:-1])
    return f"{first} {last}"




def short_name_msurname(full_name: str) -> str:
    """Z 'Meno Priezvisko' urobí 'M. Priezvisko'."""
    if not isinstance(full_name, str):
        return ""
    parts = full_name.strip().split()
    if not parts:
        return ""
    first = parts[0]
    last = parts[-1]
    initial = (first[0] + ".") if first else ""
    return (initial + " " + last).strip()


def short_pair_names(val: str) -> str:
    """Z textu 'Meno1 Priezvisko1, Meno2 Priezvisko2' urobí 'M. Priezvisko1, M. Priezvisko2'."""
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return s
    parts = [x.strip() for x in s.split(",")]
    return ", ".join(short_name_msurname(x) if x else "" for x in parts)
def players_for_year_pairs_only(df_year: pd.DataFrame):
    """Vracia (lefties, righties) zoznamy hráčov pre daný rok – IBA z L1,L2,R1,R2."""
    left_set, right_set = set(), set()
    for _, r in df_year.iterrows():
        for col in ("L1", "L2"):
            nm = _clean_name(r.get(col))
            if nm:
                left_set.add(nm)
        for col in ("R1", "R2"):
            nm = _clean_name(r.get(col))
            if nm:
                right_set.add(nm)
    return (sorted(left_set, key=str.casefold), sorted(right_set, key=str.casefold))


def build_team_table(df_year: pd.DataFrame, players: list[str], side: str) -> pd.DataFrame:
    # Ponechané pre tabuľky v karte Turnaje (nemá vplyv na hlavnú agregáciu v Štatistikách)
    def compute_player_stats(df_year: pd.DataFrame, player: str, side: str):
        if side == 'L':
            mask_pair = ((df_year['L1'] == player) | (df_year['L2'] == player))
            body_pairs = df_year.loc[mask_pair, 'Lbody'].fillna(0).sum() if 'Lbody' in df_year.columns else 0
            matches = int(mask_pair.sum())
        else:
            mask_pair = ((df_year['R1'] == player) | (df_year['R2'] == player))
            body_pairs = df_year.loc[mask_pair, 'Rbody'].fillna(0).sum() if 'Rbody' in df_year.columns else 0
            matches = int(mask_pair.sum())
        body = float(body_pairs)
        return body, matches

    def _format_body(val: float) -> str:
        return f"{int(val)}" if float(val).is_integer() else f"{val:.1f}"

    rows = []
    for p in players:
        body, matches = compute_player_stats(df_year, p, side)
        success = f"{int(round((body / matches) * 100))} %" if matches > 0 else "0 %"
        display_name = to_firstname_first(p)
        rows.append({
            "Por.": None,
            "Hráč": display_name,
            "Body": _format_body(body),
            "Zápasy": matches,
            "Úspešnosť": success,
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df.sort_values("Hráč", key=lambda s: s.str.casefold(), inplace=True)
        df.reset_index(drop=True, inplace=True)
        df = df[["Hráč", "Body", "Zápasy", "Úspešnosť"]]
    return df


def style_team_table(df: pd.DataFrame, side: str) -> Styler:
    header_bg = "#eeeeee"
    bg = COLOR_LEFT_BG if side == 'L' else COLOR_RIGHT_BG

    styler = df.style.set_properties(**{"background-color": bg, "width": "auto"})
    cols_to_center = [c for c in df.columns if c != "Hráč"]
    if cols_to_center:
        styler = styler.set_properties(subset=cols_to_center, **{"text-align": "center"})

    styler = styler.set_table_styles([
        {"selector": "th", "props": f"font-weight:700; text-align:center; background-color:{header_bg};"}
    ])

    try:
        styler = styler.hide(axis="index")
    except Exception:
        styler = styler.hide_index()

    return styler

def style_matches_table(df: pd.DataFrame) -> Styler:
    """Styler pre tabuľku zápasov: podfarbenie podľa víťaza, centrovanie, skrytý index.
    Podporuje aj mobilné skratky stĺpcov: D/Z/F/L/R/V.
    """
    header_bg = "#eeeeee"

    day_col = "Deň" if "Deň" in df.columns else ("D" if "D" in df.columns else None)
    if day_col:
        day_clean = df[day_col].astype(str).str.strip().str.replace(r"\.$", "", regex=True)
        day_series = pd.to_numeric(day_clean, errors="coerce").astype("Int64")
        df = df.copy()
        df[day_col] = day_series

    def _row_bg(row: pd.Series):
        w = str(row.get("Víťaz", row.get("V", ""))).strip().lower()
        if w in ("lefties", "l"):
            bg = COLOR_LEFT_BG
        elif w in ("righties", "r"):
            bg = COLOR_RIGHT_BG
        else:
            bg = "inherit"
        return [f"background-color: {bg}"] * len(row)

    styler = df.style.apply(_row_bg, axis=1)

    if day_col:
        styler = styler.format(subset=[day_col], formatter=lambda v: "" if pd.isna(v) else f"{int(v)}")

    cols_to_center = [
        c for c in df.columns
        if c in ["Rok","Deň","Zápas","Formát","Lefties","Righties","Víťaz","D","Z","F","L","R","V","A/S"]
    ]
    if cols_to_center:
        styler = styler.set_properties(subset=cols_to_center, **{"text-align": "center"})

    styler = styler.set_table_styles([
        {"selector": "th", "props": f"font-weight:700; text-align:center; background-color:{header_bg};"}
    ])

    try:
        styler = styler.hide(axis="index")
    except Exception:
        styler = styler.hide_index()

    return styler
    
def style_simple_table(df: pd.DataFrame, bold_last: bool = False) -> pd.io.formats.style.Styler:
    """
    Jednoduchý styler pre sumarizačné tabuľky (Formát/Rezort/Dvojice).
    - sivé hlavičky (#eeeeee)
    - centrovanie numerických stĺpcov
    - riadok 'Spolu' podfarbený rovnako sivou + tučný
    - skrytý index
    """
    header_bg = "#eeeeee"

    sty = df.style.set_table_styles([
        {"selector": "th", "props": f"font-weight:700; text-align:center; background-color:{header_bg};"},
    ])

    # Stĺpce, ktoré nechávame zarovnané doľava (ostatné centrovať)
    left_cols: list[str] = []
    if "Formát" in df.columns:
        left_cols = ["Formát"]
    elif "Rezort" in df.columns:
        left_cols = ["Rezort"]
    elif {"Lefties", "Righties"}.issubset(df.columns):
        # tabuľka párov – necháme mená dvojíc vľavo
        left_cols = ["Lefties", "Righties"]

    center_cols = [c for c in df.columns if c not in left_cols]
    if center_cols:
        sty = sty.set_properties(subset=center_cols, **{"text-align": "center"})

    # Sivé podfarbenie riadku 'Spolu'
    def _sum_row_bg(row: pd.Series):
        is_sum = any(str(row.get(col, "")).strip() == "Spolu" for col in row.index)
        return [f"background-color:{header_bg}; font-weight:700;" if is_sum else "" for _ in row]

    sty = sty.apply(_sum_row_bg, axis=1)

    # Skryť index
    try:
        sty = sty.hide(axis="index")
    except Exception:
        sty = sty.hide_index()

    # Voliteľne tučný posledný riadok (ponechávame, môže byť už aj 'Spolu')
    if bold_last and len(df) > 0:
        last_idx = df.index[-1]
        sty = sty.set_properties(subset=pd.IndexSlice[last_idx, :], **{"font-weight": "700"})

    return sty

def get_portrait_ref(players_df: pd.DataFrame, canonical_name: str) -> str | None:
    """Vráti referenciu na portrét hráča z df_players_sheet.

    Uprednostní lokálnu cestu zo stĺpca _portrait_raw (resp. Portret/Portrét),
    prípadne použije _portrait_url.
    """
    if players_df is None or players_df.empty or not canonical_name:
        return None
    try:
        sub = players_df[players_df["Hráč"].astype(str).str.strip() == str(canonical_name).strip()]
        if sub.empty:
            return None
        row = sub.iloc[0]
        for col in ("_portrait_raw", "Portret", "Portrét", "_portrait_url"):
            if col in sub.columns:
                v = row.get(col)
                if pd.notna(v) and str(v).strip():
                    return str(v).strip()
        return None
    except Exception:
        return None

def get_portrait_url(players_df: pd.DataFrame, canonical_name: str) -> str | None:
    """Spätná kompatibilita: vráti referenciu na portrét (lokálna cesta alebo URL)."""
    return get_portrait_ref(players_df, canonical_name)

# -----------------------------
# Portréty hráčov – URL/Local path + fallback
# -----------------------------

# GITHUB_RAW_BASE = "https://raw.githubusercontent.com/Jasen77/lefties-righties/main/"

@st.cache_data(show_spinner=False)
def _url_exists(url: str, timeout: float = 3.0) -> bool:
    """
    Overí, či vzdialený súbor na URL existuje (HTTP 200).
    Používa HEAD, pri zlyhaní skúsi GET (niektoré servery HEAD nepodporujú).
    """
    if not url or not isinstance(url, str):
        return False
    u = url.strip()
    if not (u.startswith('http://') or u.startswith('https://')):
        return False
    try:
        import requests
        r = requests.head(u, allow_redirects=True, timeout=timeout)
        if r.status_code == 200:
            return True
        # fallback na GET
        r = requests.get(u, stream=True, allow_redirects=True, timeout=timeout)
        return r.status_code == 200
    except Exception:
        return False

@st.cache_data(show_spinner=False)
def _local_exists(path_str: str) -> bool:
    # Lokálny súbor: relatívne cesty berieme voči BASE_DIR (adresár app.py).
    if not path_str or not isinstance(path_str, str):
        return False
    s = path_str.strip()
    if not s:
        return False
    if s.startswith('http://') or s.startswith('https://'):
        return False
    try:
        pp = Path(s)
        if not pp.is_absolute():
            pp = BASE_DIR / pp
        return pp.exists()
    except Exception:
        return False

def resolve_portrait_ref(ref) -> str | None:
    # Lokálne portréty: Excel obsahuje iba názov súboru (napr. 'SegéňJ.jpg').
    if ref is None or (isinstance(ref, float) and pd.isna(ref)):
        return None
    s = str(ref).strip()
    if not s:
        return None

    if re.match(r'^https?://', s, flags=re.IGNORECASE):
        return None

    mm = re.search(r'\(([^)]+)\)', s)
    if mm:
        cand = mm.group(1).strip()
        if cand and not re.match(r'^https?://', cand, flags=re.IGNORECASE):
            s = cand

    if s.startswith(('/', '\\')) or re.match(r'^[A-Za-z]:\\', s):
        return None

    s = s.replace('\\', '/').lstrip('./')
    if s.lower().startswith('players/'):
        s = s.split('/', 1)[1]

    return str(PLAYERS_DIR / s) if s else None

def portrait_with_fallback(players_df: pd.DataFrame, canonical_name: str) -> str:
    # Vráti lokálnu cestu k portrétu hráča, inak ANONYM_FILE.
    ref = get_portrait_url(players_df, canonical_name)
    ref = resolve_portrait_ref(ref)

    if not ref and players_df is not None and not players_df.empty and canonical_name:
        try:
            sub = players_df[players_df['Hráč'].astype(str).str.strip() == str(canonical_name).strip()]
            if not sub.empty:
                for col in ('_portrait_raw', 'Portret', 'Portrét', '_portrait_url'):
                    if col in sub.columns:
                        v = sub.iloc[0][col]
                        ref2 = resolve_portrait_ref(v)
                        if ref2:
                            ref = ref2
                            break
        except Exception:
            pass

    if ref and _local_exists(ref):
        return ref

    return resolve_portrait_ref(ANONYM_FILE) or str(PLAYERS_DIR / Path(ANONYM_FILE).name)

# -----------------------------
# Globálny stav filtra + perzistencia (JSON) – PER-USER
# -----------------------------

# -- Identifikácia používateľa pre per-user JSON

def _current_user_id():
    """Zistí identifikátor používateľa pre názov JSON súboru.
       Pokus: Streamlit experimental_user → fallback: OS login."""
    try:
        u = getattr(st, "experimental_user", None)
        if callable(u):
            info = u()
            if isinstance(info, dict):
                for k in ("username", "email", "name", "user"):
                    if info.get(k):
                        return str(info.get(k))
    except Exception:
        pass
    try:
        import getpass
        return getpass.getuser() or "default"
    except Exception:
        return "default"

_uid = _current_user_id()
_uid_s = "".join(ch if (ch.isalnum() or ch in "._-") else "_" for ch in _uid)
FILTER_JSON_FILE = f"Filter/filter_state_{_uid_s}.json"

@dataclass
class FilterState:
    t_all: bool = True
    t_selected: list[str] = field(default_factory=list)            # "Rok - Rezort"
    t_child_map: dict[str, bool] = field(default_factory=dict)     # key -> bool
    teams: list[str] = field(default_factory=lambda: ['Lefties', 'Righties'])
    formats: list[str] = field(default_factory=lambda: ['Foursome', 'Fourball', 'Single'])


FILTER = FilterState()


def _build_tournament_items(df_tournaments: pd.DataFrame) -> list[dict]:
    tdf = df_tournaments.copy()
    if "Rok" in tdf.columns:
        tdf["Rok"] = pd.to_numeric(tdf["Rok"], errors="coerce").astype("Int64")
        tdf = tdf.sort_values("Rok", ascending=False)
    if "Rezort" not in tdf.columns:
        tdf["Rezort"] = ""
    items = []
    for i, r in tdf.iterrows():
        year = r.get("Rok")
        rezort = str(r.get("Rezort", "")).strip()
        key = f"flt_t_{int(year) if pd.notna(year) else i}"
        label = f"{int(year) if pd.notna(year) else ''} - {rezort}".strip(" -")
        items.append({"key": key, "label": label})
    return items


def update_filter_from_session() -> None:
    FILTER.t_all = st.session_state.get('flt_t_all', True)
    keys = st.session_state.get('flt_t_keys', [])
    FILTER.t_child_map = {k: st.session_state.get(k, False) for k in keys}
    FILTER.t_selected = st.session_state.get('flt_tournaments', [])
    FILTER.teams = st.session_state.get('flt_teams', [])
    FILTER.formats = st.session_state.get('flt_formats', [])


def _save_filter_to_json() -> None:
    data = {
        "version": 1,
        "t_all": st.session_state.get('flt_t_all', True),
        "t_selected_labels": st.session_state.get('flt_tournaments', []),
        "teams": st.session_state.get('flt_teams', []),
        "formats": st.session_state.get('flt_formats', []),
        # NOVÉ: vybraný hráč v detaile hráča
        "player_selected_display": st.session_state.get('player_detail_selected_display', None),
        "stats_hide_one_tournament": st.session_state.get('stats_hide_one_tournament', False),
    }
    try:
        Path(FILTER_JSON_FILE).parent.mkdir(parents=True, exist_ok=True)
        Path(FILTER_JSON_FILE).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass

def _load_filter_from_json() -> dict | None:
    p = Path(FILTER_JSON_FILE)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def bootstrap_filter_state() -> None:
    """Inicializácia filtra so spätným načítaním z JSON.
    Pozn.: Streamlit môže pri zmene kódu/refreshi zachovať session_state,
    preto reloadujeme aj vtedy, keď sa zmení mtime JSON súboru.
    """
    # Zisti mtime JSON (ak existuje)
    p_json = Path(FILTER_JSON_FILE)
    json_mtime = p_json.stat().st_mtime if p_json.exists() else None
    last_mtime = st.session_state.get('flt_json_mtime')
    need_reload = (not st.session_state.get('flt_bootstrapped')) or (json_mtime is not None and json_mtime != last_mtime)

    # Ak netreba reload, iba zosynchronizuj globálny FILTER a skonči
    if not need_reload:
        update_filter_from_session()
        return

    # 1) Vytvor položky turnajov a ulož ich kľúče (na to sa viažu checkboxy).
    items = _build_tournament_items(df_tournaments)
    st.session_state['flt_t_keys'] = [it['key'] for it in items]
    st.session_state.setdefault('flt_json_mtime', None)

    # 2) Defaultné nastavenia (prvý štart bez JSON)
    for it in items:
        st.session_state.setdefault(it['key'], True)
    st.session_state.setdefault('flt_t_all', True)
    st.session_state['flt_tournaments'] = [it['label'] for it in items]

    st.session_state.setdefault('flt_team_lefties', True)
    st.session_state.setdefault('flt_team_righties', True)
    st.session_state['flt_teams'] = ['Lefties', 'Righties']

    st.session_state.setdefault('flt_fmt_foursome', True)
    st.session_state.setdefault('flt_fmt_fourball', True)
    st.session_state.setdefault('flt_fmt_single', True)
    st.session_state['flt_formats'] = ['Foursome', 'Fourball', 'Single']

    # kľúč pre Detail hráča (môže byť None, ale musí existovať)
    st.session_state.setdefault('player_detail_selected_display', None)
    st.session_state.setdefault('stats_hide_one_tournament', False)

    # 3) Načítaj JSON (ak existuje) a aplikuj hodnoty
    saved = _load_filter_from_json()
    if saved:
        # načítaj master stav z JSON (ak chýba, dopočítaj neskôr)
        st.session_state['flt_t_all'] = bool(saved.get('t_all', False))
        labels_sel = list(saved.get('t_selected_labels', []))
        years_sel = set()
        for lbl in labels_sel:
            try:
                years_sel.add(int(str(lbl).split(' - ')[0].strip()))
            except Exception:
                pass
        for it in items:
            if it['label'] in labels_sel:
                st.session_state[it['key']] = True
            else:
                try:
                    y_it = int(str(it['label']).split(' - ')[0].strip())
                    st.session_state[it['key']] = (y_it in years_sel) if years_sel else False
                except Exception:
                    st.session_state[it['key']] = False

        # ak je master v JSON True, označ všetky deti; inak dopočítaj z detí
        if st.session_state.get('flt_t_all', False):
            for k in st.session_state.get('flt_t_keys', []):
                st.session_state[k] = True
        st.session_state['flt_t_all'] = all(st.session_state.get(k, False) for k in st.session_state.get('flt_t_keys', []))
        st.session_state['flt_tournaments'] = [it['label'] for it in items if st.session_state[it['key']]]

        teams = saved.get('teams', ['Lefties', 'Righties'])
        st.session_state['flt_team_lefties']  = ('Lefties'  in teams)
        st.session_state['flt_team_righties'] = ('Righties' in teams)
        st.session_state['flt_teams'] = teams

        fmts = saved.get('formats', ['Foursome', 'Fourball', 'Single'])
        st.session_state['flt_fmt_foursome'] = ('Foursome' in fmts)
        st.session_state['flt_fmt_fourball'] = ('Fourball' in fmts)
        st.session_state['flt_fmt_single']   = ('Single'   in fmts)
        st.session_state['flt_formats'] = fmts

        # posledný zvolený hráč pre Detail hráča (môže byť None)
        st.session_state['player_detail_selected_display'] = saved.get('player_selected_display', None)
        st.session_state['stats_hide_one_tournament'] = bool(saved.get('stats_hide_one_tournament', False))

    # 4) Označ, že bootstrap prebehol, zosynchronizuj cache a hneď ulož stav do JSON.
    st.session_state['flt_bootstrapped'] = True
    st.session_state['flt_json_mtime'] = json_mtime
    update_filter_from_session()
    _save_filter_to_json()


def _on_filter_change() -> None:
    items = _build_tournament_items(df_tournaments)

    selected = [it['label'] for it in items if st.session_state.get(it['key'], False)]
    st.session_state['flt_tournaments'] = selected
    st.session_state['flt_t_all'] = all(st.session_state.get(k, False) for k in st.session_state.get('flt_t_keys', []))

    teams = []
    if st.session_state.get('flt_team_lefties'):
        teams.append('Lefties')
    if st.session_state.get('flt_team_righties'):
        teams.append('Righties')
    st.session_state['flt_teams'] = teams

    fmts = []
    if st.session_state.get('flt_fmt_foursome'):
        fmts.append('Foursome')
    if st.session_state.get('flt_fmt_fourball'):
        fmts.append('Fourball')
    if st.session_state.get('flt_fmt_single'):
        fmts.append('Single')
    st.session_state['flt_formats'] = fmts

    update_filter_from_session()
    _save_filter_to_json()


def _toggle_all_tournaments() -> None:
    val = st.session_state.get('flt_t_all', True)
    for k in st.session_state.get('flt_t_keys', []):
        st.session_state[k] = val
    _on_filter_change()

def _on_player_select_change() -> None:
    # iba persist – UI si prečíta st.session_state
    _save_filter_to_json()
    
# -- Bootstrap pri štarte
bootstrap_filter_state()


# -----------------------------
# NOVÉ jadro Štatistík podľa pravidiel 1–6
# -----------------------------

def build_player_team_map(df_all: pd.DataFrame) -> dict[str, str]:
    """Zaradenie hráčov do tímov podľa výskytu v L1/L2 (Lefties) a R1/R2 (Righties).
       Pri výskyte v oboch: použijeme vyšší počet; pri rovnosti preferuj Lefties."""
    cntL, cntR = {}, {}
    for col in ("L1", "L2"):
        if col in df_all.columns:
            for nm in df_all[col].dropna().astype(str).str.strip():
                if nm:
                    cntL[nm] = cntL.get(nm, 0) + 1
    for col in ("R1", "R2"):
        if col in df_all.columns:
            for nm in df_all[col].dropna().astype(str).str.strip():
                if nm:
                    cntR[nm] = cntR.get(nm, 0) + 1

    players = set(cntL) | set(cntR)
    team = {}
    for p in players:
        l, r = cntL.get(p, 0), cntR.get(p, 0)
        if l > r:
            team[p] = "Lefties"
        elif r > l:
            team[p] = "Righties"
        else:
            team[p] = "Lefties" if l > 0 else ("Righties" if r > 0 else "Lefties")
    return team


def compute_stats_for_filtered(
    df_matches: pd.DataFrame,
    sel_years: list[int],
    sel_formats: set[str],
    sel_teams: set[str],
    team_map: dict[str, str],
):
    """Prejde vyfiltrované zápasy a spočíta body + zápasy pre hráčov podľa strán.
       LEFT hráči berú Lbody; RIGHT hráči berú Rbody. Formát = stĺpec "Formát"."""
    from collections import defaultdict

    # Guard: ak nie je vybraný žiaden formát, nepočítaj nič
    if sel_formats is not None and len(sel_formats) == 0:
        return [], []
    # Guard: ak nie je vybraný žiaden turnaj/rok (prázdny výber), nepočítaj nič
    # Pozn.: prázdny zoznam rokov znamená 'nič vybraté' (nie 'všetko')
    if sel_years is not None and len(sel_years) == 0:
        return [], []

    df_y = df_matches.copy()
    if sel_years:
        df_y = df_y[df_y["Rok"].isin(sel_years)]
    if sel_formats:
        df_y = df_y[df_y["Formát"].isin(sel_formats)]

    FMT_KEYS = ("Foursome", "Fourball", "Single")

    def _empty_bucket():
        return {
            "Team": "",
            "Foursome": {"pts": 0.0, "cnt": 0},
            "Fourball": {"pts": 0.0, "cnt": 0},
            "Single": {"pts": 0.0, "cnt": 0},
        }

    stats = defaultdict(_empty_bucket)

    for _, row in df_y.iterrows():
        fmt = str(row.get("Formát", "")).strip()
        if fmt not in FMT_KEYS:
            continue

        # hráči na ľavej a pravej strane
        left_names, right_names = [], []
        for col in ("L1", "L2"):
            if col in df_y.columns:
                nm = _clean_name(row.get(col))
                if nm:
                    left_names.append(nm)
        for col in ("R1", "R2"):
            if col in df_y.columns:
                nm = _clean_name(row.get(col))
                if nm:
                    right_names.append(nm)

        # body riadku
        lbody = row.get("Lbody", 0)
        rbody = row.get("Rbody", 0)
        try:
            lbody = float(lbody) if lbody is not None else 0.0
        except Exception:
            lbody = 0.0
        try:
            rbody = float(rbody) if rbody is not None else 0.0
        except Exception:
            rbody = 0.0

        # ľavá strana -> Lbody
        for p in left_names:
            p_team = team_map.get(p, "Lefties")
            if sel_teams and (p_team not in sel_teams):
                continue
            b = stats[p]
            b["Team"] = p_team
            b[fmt]["pts"] += lbody
            b[fmt]["cnt"] += 1

        # pravá strana -> Rbody
        for p in right_names:
            p_team = team_map.get(p, "Righties")
            if sel_teams and (p_team not in sel_teams):
                continue
            b = stats[p]
            b["Team"] = p_team
            b[fmt]["pts"] += rbody
            b[fmt]["cnt"] += 1

    def _fmt_points(x: float) -> str:
        return f"{int(x)}" if float(x).is_integer() else f"{x:.1f}"

    def _pct(points_sum: float, cnt: int) -> int:
        return int(round((points_sum / cnt) * 100)) if cnt else 0

    rows_disp, rows_num = [], []
    for p, b in stats.items():
        team = b["Team"] or team_map.get(p, "Lefties")
        fs_pts, fs_cnt = b["Foursome"]["pts"], b["Foursome"]["cnt"]
        fb_pts, fb_cnt = b["Fourball"]["pts"], b["Fourball"]["cnt"]
        si_pts, si_cnt = b["Single"]["pts"], b["Single"]["cnt"]
        total_pts = fs_pts + fb_pts + si_pts
        total_cnt = fs_cnt + fb_cnt + si_cnt

        rows_disp.append({
            'Hráč': to_firstname_first(p),
            'Team': team,
            'Foursome Body': _fmt_points(fs_pts),
            'Foursome Zápasy': fs_cnt,
            'Foursome Úsp.': f"{_pct(fs_pts, fs_cnt)} %",
            'Fourball Body': _fmt_points(fb_pts),
            'Fourball Zápasy': fb_cnt,
            'Fourball Úsp.': f"{_pct(fb_pts, fb_cnt)} %",
            'Single Body': _fmt_points(si_pts),
            'Single Zápasy': si_cnt,
            'Single Úsp.': f"{_pct(si_pts, si_cnt)} %",
            'Spolu Body': _fmt_points(total_pts),
            'Spolu Zápasy': total_cnt,
            'Spolu Úsp.': f"{_pct(total_pts, total_cnt)} %",
        })

        rows_num.append({
            'Hráč': to_firstname_first(p),
            'Team': team,
            'Foursome Body': float(fs_pts),
            'Foursome Zápasy': int(fs_cnt),
            'Foursome Úsp.': _pct(fs_pts, fs_cnt),
            'Fourball Body': float(fb_pts),
            'Fourball Zápasy': int(fb_cnt),
            'Fourball Úsp.': _pct(fb_pts, fb_cnt),
            'Single Body': float(si_pts),
            'Single Zápasy': int(si_cnt),
            'Single Úsp.': _pct(si_pts, si_cnt),
            'Spolu Body': float(total_pts),
            'Spolu Zápasy': int(total_cnt),
            'Spolu Úsp.': _pct(total_pts, total_cnt),
        })

    return rows_disp, rows_num


def build_player_years_count_display(df_all: pd.DataFrame) -> dict[str, int]:
    """
    Vráti mapu: 'Meno Priezvisko' (display) -> počet unikátnych ročníkov, v ktorých hráč hral
    (globálne naprieč všetkými dátami, nezávisle od filtra).
    """
    if df_all is None or df_all.empty:
        return {}

    tmp = df_all.copy()
    if "Rok" not in tmp.columns:
        return {}

    # Rok ako číslo (Int64), aby sa dal bezpečne používať v množinách
    tmp["Rok"] = pd.to_numeric(tmp["Rok"], errors="coerce").astype("Int64")

    years_by_player: dict[str, set[int]] = {}

    for col in ("L1", "L2", "R1", "R2"):
        if col not in tmp.columns:
            continue
        for _, r in tmp[["Rok", col]].dropna().iterrows():
            y = r["Rok"]
            nm = str(r[col]).strip()
            if pd.isna(y) or not nm:
                continue

            # v štatistikách používate to_firstname_first() (Meno Priezvisko),
            # tak rovnaký tvar použijeme aj tu, aby sedelo filtrovanie na df_disp['Hráč']
            disp = to_firstname_first(nm)

            years_by_player.setdefault(disp, set()).add(int(y))

    return {p: len(ys) for p, ys in years_by_player.items()}
    
  

  
    
# =============================
# UI – Tabs: Turnaje | Štatistiky | Detail hráča | Filter
# =============================

tab_turnaje, tab_stats, tab_player, tab_filter = st.tabs(["Turnaje", "Štatistiky", "Detail hráča", "Filter"])




# *****************************
# Štatistiky
# *****************************
with tab_stats:
    st.subheader("Štatistiky")

    # -- Súhrn aktuálneho filtra (len riadky; prvý riadok začína **Turnaje:**)
    def _filter_summary_from_global() -> str:
        if FILTER.t_all:
            t_str = "všetky turnaje"
        else:
            years = []
            for lbl in FILTER.t_selected:
                try:
                    y = int(str(lbl).split(' - ')[0].strip())
                    years.append(str(y))
                except Exception:
                    pass
            t_str = ", ".join(sorted(set(years))) if years else "—"
        teams_str = ", ".join(FILTER.teams) if FILTER.teams else "—"
        fmts_str = ", ".join(FILTER.formats) if FILTER.formats else "—"
        return (
            f"Filter:  \n"
            f"**Turnaje:** {t_str}  \n"
            f"**Tímy:** {teams_str}  \n"
            f"**Formáty:** {fmts_str}"
        )
    st.markdown(_filter_summary_from_global())

    # --- Roky z FILTER.t_selected ---
    years_list = []
    for lbl in FILTER.t_selected:
        try:
            years_list.append(int(str(lbl).split(' - ')[0].strip()))
        except Exception:
            pass
    years_list = sorted(set(years_list))

    # --- Filtre ---
    sel_years = years_list
    sel_formats = set(FILTER.formats)
    sel_teams = set(FILTER.teams)

    # --- Team mapa a prepočet ---
    player_team_map = build_player_team_map(df_matches)
    rows_disp, rows_num = compute_stats_for_filtered(
        df_matches=df_matches,
        sel_years=sel_years,
        sel_formats=sel_formats,
        sel_teams=sel_teams,
        team_map=player_team_map,
    )

    import pandas as pd
    df_disp = pd.DataFrame(rows_disp)
    df_num = pd.DataFrame(rows_num)

    # --- Globálne: ročníky účasti hráča (nezávisle od filtra) ---
    player_years_count = build_player_years_count_display(df_matches)
    one_year_players = {p for p, cnt in player_years_count.items() if cnt == 1}

    # --- Počet "skrytých" pre aktuálny filter (počítame vždy, aj keď checkbox nie je zaškrtnutý) ---
    hidden_now = 0
    if not df_disp.empty and "Hráč" in df_disp.columns:
        hidden_now = int(df_disp["Hráč"].isin(one_year_players).sum())
    
    # --- Checkbox logika (platí len pre kartu Štatistiky) ---
    if st.session_state.get("stats_hide_one_tournament", False):
        one_year_players = {p for p, cnt in player_years_count.items() if cnt == 1}

        # koľko z hráčov, ktorí vyhovujú AKTUÁLNEMU filtru (t.j. sú v df_disp),
        # je zároveň v skupine "iba 1 ročník" (globálne)
        hidden_now = 0
        if not df_disp.empty and "Hráč" in df_disp.columns:
            hidden_now = int(df_disp["Hráč"].isin(one_year_players).sum())

        # st.caption(f"Počet skrytých hráčov pre aktuálny filter: {hidden_now}")

        # až potom reálne odfiltruj, aby sa v tabuľke nezobrazili
        if not df_disp.empty:
            df_disp = df_disp[~df_disp["Hráč"].isin(one_year_players)].copy()
        if not df_num.empty:
            df_num = df_num[~df_num["Hráč"].isin(one_year_players)].copy()
        
    if df_disp.empty:
        st.info("Pre zvolený filter nie sú k dispozícii dáta na zobrazenie.")
    else:
        # --- DYNAMICKÉ zoradenie cez SELECTBOX (namiesto buttonov) ---
        # Locale pre abecedné zoradenie podľa priezviska
        import locale

        def _set_sk_locale_once():
            if 'sk_locale_ok' in st.session_state:
                return
            ok = False
            for loc in ('sk_SK.UTF-8', 'sk_SK', 'Slovak_Slovakia.1250', 'cs_CZ.UTF-8', 'cs_CZ'):
                try:
                    locale.setlocale(locale.LC_COLLATE, loc)
                    ok = True
                    break
                except Exception:
                    pass
            st.session_state['sk_locale_ok'] = ok

        def _sk_xfrm(s: str) -> str:
            return locale.strxfrm(s) if st.session_state.get('sk_locale_ok') else s.casefold()

        def _surname(full_name: str) -> str:
            if not isinstance(full_name, str):
                return ''
            parts = full_name.strip().split()
            return parts[-1] if parts else ''

        _set_sk_locale_once()

        # Priprav preklad z formátov na tokeny
        FORMAT_ORDER = [('Foursome', 'Fs'), ('Fourball', 'Fb'), ('Single', 'Si')]
        included = [(fmt, tag) for fmt, tag in FORMAT_ORDER if fmt in sel_formats]

        _sort_map = {}
        for fmt, tag in included:
            _sort_map[f'{tag}B'] = f'{fmt} Body'
            _sort_map[f'{tag}Z'] = f'{fmt} Zápasy'
            _sort_map[f'{tag}Ú'] = f'{fmt} Úsp.'


        # Zobrazované voľby v selectboxe (default = SpÚ)
        options = [
            ("SpB", "Spolu – Body"),
            ("SpZ", "Spolu – Zápasy"),
            ("SpÚ", "Spolu – Úspešnosť"),  # implicitná voľba
        ]
        for fmt, tag in included:  # included je už v poradí Foursome, Fourball, Single
            options.append((f"{tag}B", f"{fmt} – Body"))
            options.append((f"{tag}Z", f"{fmt} – Zápasy"))
            options.append((f"{tag}Ú", f"{fmt} – Úspešnosť"))
        options.append(("Abc", "Abecedne podľa priezviska"))

        # Predvolená/aktuálna voľba
        default_token = "SpÚ"
        if 'stats_sort' in st.session_state:
            cur_key, _ = st.session_state['stats_sort']
            if cur_key == 'ABC':
                default_token = 'Abc'
            elif cur_key in ('Spolu Body', 'Spolu Zápasy', 'Spolu Úsp.'):
                default_token = {'Spolu Body': 'SpB', 'Spolu Zápasy': 'SpZ', 'Spolu Úsp.': 'SpÚ'}[cur_key]
            else:
                for token, colname in _sort_map.items():
                    if colname == cur_key:
                        default_token = token
                        break

        display_labels = [label for _, label in options]
        tokens = [tok for tok, _ in options]
        try:
            default_index = tokens.index(default_token)
        except ValueError:
            default_index = tokens.index("SpÚ")

        # --- CSS: selectbox + tabuľka v spoločnom wrapperi so zdieľanou šírkou ---
        STATS_FIT_CSS = r"""
        /* Kontajner pre selectbox + tabuľku: šírka = šírka tabuľky */
        .stats-fit {
          display: inline-block;
          max-width: 100%;
          vertical-align: top;
        }
        /* tabuľka nech má prirodzenú šírku (určí šírku .stats-fit) */
        .stats-fit table { width: auto !important; }
        /* selectbox sa natiahne na 100 % šírky .stats-fit (teda = šírka tabuľky) */
        .stats-fit [data-testid="stSelectbox"],
        .stats-fit [data-testid="stSelectbox"] > div,
        .stats-fit [data-baseweb="select"],
        .stats-fit [data-baseweb="select"] > div {
          width: 100% !important;
          min-width: 0 !important;
        }
        .stats-fit [data-testid="stSelectbox"] { margin-bottom: 0.5rem; }
        """
        st.markdown(f"<style>{STATS_FIT_CSS}</style>", unsafe_allow_html=True)

        # --- Začiatok spoločného wrappera (selectbox + tabuľka) ---
        st.markdown('<div class="stats-fit">', unsafe_allow_html=True)

        hide_one_tournament = st.checkbox(
            f"Vynechať {hidden_now} hráčov s účasťou iba na jednom turnaji",
            key="stats_hide_one_tournament",
            value=st.session_state.get('stats_hide_one_tournament', False),
            on_change=_save_filter_to_json,
            help="Ak je zapnuté, v Štatistikách sa skryjú hráči, ktorí sa v celej histórii zúčastnili iba 1 ročníka."
        )
        
        if hide_one_tournament:
            if not df_disp.empty:
                df_disp = df_disp[~df_disp["Hráč"].isin(one_year_players)].copy()
            if not df_num.empty:
                df_num = df_num[~df_num["Hráč"].isin(one_year_players)].copy()
        
        selected_label = st.selectbox(
            "Zoradenie tabuľky",
            display_labels,
            index=default_index,
            key="stats_sort_select",
            help=(
                "Vyber poradie zobrazenia: Abecedne podľa priezviska, alebo podľa Body/Zápasy/Úspešnosť "
                "pre jednotlivé formáty (Foursome/Fourball/Single) či súhrn 'Spolu'."
            ),
        )
        selected_token = tokens[display_labels.index(selected_label)]
        if selected_token == 'Abc':
            st.session_state['stats_sort'] = ('ABC', True)  # abecedne
        elif selected_token in ('SpB', 'SpZ', 'SpÚ'):
            name = {'SpB': 'Spolu Body', 'SpZ': 'Spolu Zápasy', 'SpÚ': 'Spolu Úsp.'}[selected_token]
            st.session_state['stats_sort'] = (name, False)  # čísla/percentá zostupne
        else:
            # FsB/FsZ/FsÚ/FbB/... premapujeme cez _sort_map na interný stĺpec
            st.session_state['stats_sort'] = (_sort_map[selected_token], False)

        # --- Aplikácia zoradenia na df_disp/df_num (ako doteraz)
        allowed_sort_cols = {'ABC', 'Spolu Body', 'Spolu Zápasy', 'Spolu Úsp.'}
        allowed_sort_cols |= set(_sort_map.values())
        if ('stats_sort' not in st.session_state) or (st.session_state['stats_sort'][0] not in allowed_sort_cols):
            st.session_state['stats_sort'] = ('Spolu Úsp.', False)

        sort_key, sort_asc = st.session_state['stats_sort']
        if sort_key == 'ABC':
            df_disp['_sort1'] = df_disp['Hráč'].apply(lambda x: _sk_xfrm(_surname(x)))
            df_disp['_sort2'] = df_disp['Hráč'].apply(lambda x: _sk_xfrm(x))
            df_disp.sort_values(by=['_sort1', '_sort2'], ascending=[True, True], inplace=True)
            df_disp.drop(columns=['_sort1', '_sort2'], inplace=True)
        else:
            df_disp['_sort_val'] = df_num[sort_key]
            df_disp['_sort_name'] = df_disp['Hráč'].apply(lambda x: _sk_xfrm(x))
            df_disp.sort_values(by=['_sort_val', '_sort_name'], ascending=[sort_asc, True], inplace=True)
            df_disp.drop(columns=['_sort_val', '_sort_name'], inplace=True)

        # --- Poradie stĺpcov podľa vybraných formátov (desktop vs mobil)
        # Mobil: iba P, Hráč (M. Priezvisko), T (L/R) + 1 sekcia podľa zvoleného zoradenia (Foursome/Fourball/Single/Spolu)
        # Pre export: vždy desktop reprezentácia (aj na mobile)
        df_stats_export_source = df_disp.copy()

        is_mobile = (globals().get('_device_type') == 'mobil')

        def _short_name(full_name: str) -> str:
            if not isinstance(full_name, str):
                return ''
            parts = full_name.strip().split()
            if not parts:
                return ''
            first = parts[0]
            last = parts[-1]
            initial = (first[0] + '.') if first else ''
            return (initial + ' ' + last).strip()

        def _team_short(v: str) -> str:
            s = str(v).strip()
            return 'L' if s == 'Lefties' else ('R' if s == 'Righties' else s)

        if is_mobile:
            # sekcia podľa sort_key; pri abecednom zoradení vždy Spolu
            sec = 'Spolu'
            if sort_key != 'ABC':
                head = str(sort_key).split(' ', 1)[0]
                if head in ('Foursome', 'Fourball', 'Single') and head in sel_formats:
                    sec = head

            df_disp = df_disp.copy()
            if 'Hráč' in df_disp.columns:
                df_disp['Hráč'] = df_disp['Hráč'].apply(_short_name)
            if 'Team' in df_disp.columns:
                df_disp['Team'] = df_disp['Team'].apply(_team_short)

            if 'Por.' in df_disp.columns:
                df_disp['Por.'] = range(1, len(df_disp) + 1)
            else:
                df_disp.insert(0, 'Por.', range(1, len(df_disp) + 1))

            if sec in ('Foursome', 'Fourball', 'Single'):
                flat_order = ['Por.', 'Hráč', 'Team', sec + ' Body', sec + ' Zápasy', sec + ' Úsp.']
                col_tuples = [('', 'P'), ('', 'Hráč'), ('', 'T'), (sec, 'B'), (sec, 'Z'), (sec, 'Ú')]
            else:
                flat_order = ['Por.', 'Hráč', 'Team', 'Spolu Body', 'Spolu Zápasy', 'Spolu Úsp.']
                col_tuples = [('', 'P'), ('', 'Hráč'), ('', 'T'), ('Spolu', 'B'), ('Spolu', 'Z'), ('Spolu', 'Ú')]

            df_disp = df_disp[flat_order].copy()
            df_disp.columns = pd.MultiIndex.from_tuples(col_tuples)

        else:
            flat_order = ['Por.', 'Hráč', 'Team']
            for fmt, _ in included:
                flat_order += [fmt + ' Body', fmt + ' Zápasy', fmt + ' Úsp.']
            flat_order += ['Spolu Body', 'Spolu Zápasy', 'Spolu Úsp.']

            if 'Por.' in df_disp.columns:
                df_disp['Por.'] = range(1, len(df_disp) + 1)
            else:
                df_disp.insert(0, 'Por.', range(1, len(df_disp) + 1))
            df_disp = df_disp[flat_order]

            # --- MultiIndex hlavičky (vizuálne skupiny stĺpcov)
            col_tuples = [('', 'Por.'), ('', 'Hráč'), ('', 'Team')]
            for fmt, _ in included:
                col_tuples += [(fmt, 'Body'), (fmt, 'Zápasy'), (fmt, 'Úsp.')]
            col_tuples += [('Spolu', 'Body'), ('Spolu', 'Zápasy'), ('Spolu', 'Úsp.')]
            df_disp.columns = pd.MultiIndex.from_tuples(col_tuples)

        def _col_tuple_for_sort_key(sk: str):
            # Mobil: hlavičky sú skrátené (P/T, B/Z/Ú); Desktop: pôvodné
            is_mobile_local = (globals().get('_device_type') == 'mobil')
            if sk == 'ABC':
                return ('', 'Hráč')
            if is_mobile_local:
                if sk in ('Spolu Body', 'Spolu Zápasy', 'Spolu Úsp.'):
                    return ('Spolu', {'Body':'B','Zápasy':'Z','Úsp.':'Ú'}[sk.split()[-1]])
                try:
                    fmt, metric = sk.split(' ', 1)
                    if fmt in ('Foursome','Fourball','Single'):
                        return (fmt, {'Body':'B','Zápasy':'Z','Úsp.':'Ú'}.get(metric, metric))
                except Exception:
                    return None
                return None
            else:
                if sk in ('Spolu Body', 'Spolu Zápasy', 'Spolu Úsp.'):
                    return ('Spolu', sk.split()[-1])
                try:
                    fmt, metric = sk.split(' ', 1)
                    return (fmt, metric)
                except Exception:
                    return None
        col_to_bold = _col_tuple_for_sort_key(sort_key)

        # --- Styler pre tabuľku Štatistiky
        def style_stats_table(df: pd.DataFrame, highlight_col=None) -> Styler:
            header_bg = "#eeeeee"
            styler = df.style.set_table_styles([
                {"selector": "th", "props": f"font-weight:700; text-align:center; background-color:{header_bg};"},
                {"selector": "th.col_heading", "props": f"font-weight:700; text-align:center; background-color:{header_bg};"},
                {"selector": "th.col_heading.level0", "props": f"font-weight:700; text-align:center; background-color:{header_bg};"},
                {"selector": "th.col_heading.level1", "props": f"font-weight:700; text-align:center; background-color:{header_bg};"},
            ])
            cols_center = [c for c in df.columns if c != ('', 'Hráč')]
            if cols_center:
                styler = styler.set_properties(subset=cols_center, **{"text-align": "center"})

            def _row_bg(row):
                team = str(row.get(('', 'Team'), row.get(('', 'T'), ''))).strip()
                if team in ('Lefties','L'):
                    bg = COLOR_LEFT_BG
                elif team in ('Righties','R'):
                    bg = COLOR_RIGHT_BG
                else:
                    bg = 'inherit'
                return [f'background-color: {bg}'] * len(row)

            styler = styler.apply(_row_bg, axis=1)
            if highlight_col and highlight_col in df.columns:
                styler = styler.set_properties(
                    subset=(slice(None), [highlight_col]),
                    **{"font-weight": "700", "font-size": "1.05rem"}
                )
            try:
                styler = styler.hide(axis='index')
            except Exception:
                styler = styler.hide_index()
            return styler

        sty = style_stats_table(df_disp, highlight_col=col_to_bold)
        html = sty.to_html()
        st.markdown(html, unsafe_allow_html=True)

        # --- Koniec spoločného wrappera (select + tabuľka)
        st.markdown('</div>', unsafe_allow_html=True)

        # --- Export Štatistík + Filter do Excelu ---
        try:
            # 1) Splosknutie MultiIndex stĺpcov
            def _flatten_stats_columns(cols) -> list[str]:
                flat = []
                for col in cols:
                    if isinstance(col, tuple) and len(col) == 2:
                        a, b = col
                        flat.append(b if (a == '' or a is None) else f"{a} {b}")
                    else:
                        flat.append(str(col))
                return flat

            df_stats_export = df_stats_export_source.copy() if 'df_stats_export_source' in locals() else df_disp.copy()
            if isinstance(df_stats_export.columns, pd.MultiIndex):
                df_stats_export.columns = _flatten_stats_columns(df_stats_export.columns)

            # 2) Hárok "Filter" – vypíš všetky vybrané položky
            rows_filter = []
            if getattr(FILTER, "t_all", False):
                rows_filter.append({"Kategória": "Turnaje", "Hodnota": "všetky turnaje"})
            else:
                for lbl in getattr(FILTER, "t_selected", []):
                    rows_filter.append({"Kategória": "Turnaj", "Hodnota": str(lbl)})
            teams_val = ", ".join(getattr(FILTER, "teams", [])) if getattr(FILTER, "teams", []) else "—"
            fmts_val  = ", ".join(getattr(FILTER, "formats", [])) if getattr(FILTER, "formats", []) else "—"
            rows_filter.append({"Kategória": "Tímy", "Hodnota": teams_val})
            rows_filter.append({"Kategória": "Formáty", "Hodnota": fmts_val})
            df_filter_export = pd.DataFrame(rows_filter, columns=["Kategória", "Hodnota"])

            # 3) Zápis do Excelu + centrovanie + autofit
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment
            from datetime import datetime
            import io, re

            def _write_sheet_auto_fit(writer, df: pd.DataFrame, sheet_name: str):
                df_to_save = df.copy() if (df is not None and not df.empty) else pd.DataFrame()
                df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                # Centrovanie buniek
                align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).alignment = align_center
                # Autofit šírky
                if df_to_save.empty:
                    for col_idx in range(1, max(1, ws.max_column) + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 18
                else:
                    for col_idx, col_name in enumerate(df_to_save.columns, start=1):
                        series = df_to_save[col_name].astype(str).fillna("")
                        max_len = max([len(str(col_name))] + series.map(len).tolist())
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

            timestamp = datetime.now().strftime("%Y.%m.%d-%H.%M.%S")
            xlsx_name = f"L&R - Štatistiky ({timestamp}).xlsx"
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                _write_sheet_auto_fit(writer, df_stats_export, "Štatistiky")
                _write_sheet_auto_fit(writer, df_filter_export, "Filter")

            st.download_button(
                label=f"⬇️ Export do Excelu ({xlsx_name})",
                data=buffer.getvalue(),
                file_name=xlsx_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_xlsx_stats",
            )

        except Exception as _ex:
            st.warning(f"Export Štatistík do Excelu sa nepodaril: {type(_ex).__name__}: {_ex}")        
        



# *****************************
# Detail hráča
# *****************************
with tab_player:
    st.subheader("Detail hráča")

    # -- Súhrn aktuálneho filtra (len riadky; prvý riadok začína **Turnaje:**)
    def _filter_summary_from_global() -> str:
        if FILTER.t_all:
            t_str = "všetky turnaje"
        else:
            years = []
            for lbl in FILTER.t_selected:
                try:
                    y = int(str(lbl).split(' - ')[0].strip())
                    years.append(str(y))
                except Exception:
                    pass
            t_str = ", ".join(sorted(set(years))) if years else "—"
        teams_str = ", ".join(FILTER.teams) if FILTER.teams else "—"
        fmts_str = ", ".join(FILTER.formats) if FILTER.formats else "—"
        return (
            f"Filter:  \n"
            f"**Turnaje:** {t_str}  \n"
            f"**Tímy:** {teams_str}  \n"
            f"**Formáty:** {fmts_str}"
        )
    st.markdown(_filter_summary_from_global())

    # -- Z aktívneho FILTRA načítaj roky, tímy a formáty
    years_list = []
    for lbl in FILTER.t_selected:
        try:
            years_list.append(int(str(lbl).split(' - ')[0].strip()))
        except Exception:
            pass
    sel_years = sorted(set(years_list))
    sel_teams = set(FILTER.teams or [])            # {'Lefties','Righties'} alebo prázdne
    sel_formats = set(FILTER.formats or [])        # {'Foursome','Fourball','Single'} alebo prázdne

    # --- 1) Najprv zostav hráčov len podľa rokov+ tímov (NEZÁVISLE od formátov) ---
    df_players_src = df_matches.copy()
    if sel_years:
        df_players_src = df_players_src[df_players_src["Rok"].isin(sel_years)]

    players_set = set()
    if "Lefties" in sel_teams:
        for col in ("L1", "L2"):
            if col in df_players_src.columns:
                players_set.update(n for n in df_players_src[col].dropna().astype(str).str.strip() if n)

    if "Righties" in sel_teams:
        for col in ("R1", "R2"):
            if col in df_players_src.columns:
                players_set.update(n for n in df_players_src[col].dropna().astype(str).str.strip() if n)

    players_sorted = sorted(players_set, key=str.casefold)

    if not players_sorted:
        st.info("Pre zvolenú kombináciu **Tímy** a **Turnaje** nie je k dispozícii žiadny hráč. Uprav výber vo **Filtri**.")
    else:
        # --- 2) Combobox sa vykreslí VŽDY, keď existujú hráči ---
        display_names = [to_firstname_first(p) for p in players_sorted]
        display_to_canon = dict(zip(display_names, players_sorted))
        SEL_KEY = "player_detail_selected_display"

        cur = st.session_state.get(SEL_KEY)
        if cur not in display_names:
            st.session_state[SEL_KEY] = display_names[0]

        st.selectbox(
            "Vyber hráča",
            display_names,
            key=SEL_KEY,
            on_change=_save_filter_to_json,
        )

        selected_display = st.session_state[SEL_KEY]
        selected_canonical = display_to_canon[selected_display]

        # --- Team hráča urč vždy hneď po výbere (aby existoval aj keď nie sú zvolené formáty) ---
        player_team_map = build_player_team_map(df_matches)
        player_team = player_team_map.get(selected_canonical, "Lefties")
        
        # --- 3) AŽ TERAZ rieš formáty: ak nie sú, zobraz hlášku a NEPOKRAČUJ v detailoch ---
        if not sel_formats:
            st.info("Nie je zvolený žiadny **Formát hry**. Zapni aspoň jeden v karte **Filter**.")
        else:
            # sem patrí VÁŠ PÔVODNÝ VÝPOČET DETAILU (df_output_src, tabuľky, portrét, zápasy...)
            df_output_src = df_matches.copy()
            if sel_years:
                df_output_src = df_output_src[df_output_src["Rok"].isin(sel_years)]
            df_output_src = df_output_src[df_output_src["Formát"].isin(sel_formats)]

            # ... a potom pokračuje váš kód:
            player_team_map = build_player_team_map(df_matches)
            player_team = player_team_map.get(selected_canonical, "Lefties")
            team_badge_bg = COLOR_LEFT_BG if player_team == "Lefties" else COLOR_RIGHT_BG

            st.markdown(
                f"""
                <div style='font-size:2rem; font-weight:800; margin: 8px 0 4px;'>{selected_display}</div>
                <div style='display:inline-block; padding:4px 10px; border-radius:999px;
                            background:{team_badge_bg}; font-weight:700; margin-bottom:12px;'>
                    Team {player_team}
                </div>
                """,
                unsafe_allow_html=True
            )
            # ... zvyšok detailu hráča
        



        # -- Všetky zápasy vybraného hráča (rešpektujú ROČNÍKY + FORMÁTY)
        # Guard: df_output_src musí existovať vždy (aj pri prázdnych formátoch)
        if 'df_output_src' not in locals():
            df_output_src = df_matches.copy().iloc[0:0]
        mask_player = False
        for col in ("L1", "L2", "R1", "R2"):
            if col in df_output_src.columns:
                mask_player = mask_player | (df_output_src[col].astype(str).str.strip() == selected_canonical)
        df_player = df_output_src.loc[mask_player].copy()

        # -- Výpočet bodov pre hráča po riadkoch (ak je vľavo -> Lbody, ak vpravo -> Rbody)
        def _points_for_row(row) -> float:
            try:
                is_left = (str(row.get("L1","")).strip() == selected_canonical) or (str(row.get("L2","")).strip() == selected_canonical)
                is_right = (str(row.get("R1","")).strip() == selected_canonical) or (str(row.get("R2","")).strip() == selected_canonical)
                lb = float(row.get("Lbody", 0) or 0)
                rb = float(row.get("Rbody", 0) or 0)
                if is_left and not is_right:
                    return lb
                if is_right and not is_left:
                    return rb
                return lb if player_team == "Lefties" else rb
            except Exception:
                return 0.0

        if not df_player.empty:
            df_player["_points"] = df_player.apply(_points_for_row, axis=1).astype(float)
        else:
            df_player["_points"] = []

        # -- SUMÁR CELOKOM podľa formátu (Foursome/Fourball/Single) + riadok Spolu
        ORDER = ["Foursome", "Fourball", "Single"]
        formats_in_scope = [f for f in ORDER if (not sel_formats) or (f in sel_formats)]

        def _fmt_pts(x: float) -> str:
            return f"{int(x)}" if float(x).is_integer() else f"{x:.1f}"

        def _pct(pts: float, cnt: int) -> int:
            return int(round((pts / cnt) * 100)) if cnt else 0

        agg_fmt = []
        tot_pts = 0.0
        tot_cnt = 0
        for fmt in formats_in_scope:
            sub = df_player[df_player["Formát"] == fmt] if "Formát" in df_player.columns else df_player.iloc[0:0]
            pts = float(sub["_points"].sum()) if not sub.empty else 0.0
            cnt = int(len(sub))
            agg_fmt.append({"Formát": fmt, "Body": _fmt_pts(pts), "Zápasy": cnt, "Úspešnosť": f"{_pct(pts, cnt)} %"})
            tot_pts += pts
            tot_cnt += cnt

        agg_fmt.append({"Formát": "Spolu", "Body": _fmt_pts(tot_pts), "Zápasy": tot_cnt, "Úspešnosť": f"{_pct(tot_pts, tot_cnt)} %"})
        df_fmt_sum = pd.DataFrame(agg_fmt)

        # -- Portrét hráča (200x200) – lokálna cesta z Excelu + fallback na ANONYM_FILE
        portrait_ref = portrait_with_fallback(df_players_sheet, selected_canonical)
        if portrait_ref:
            st.image(portrait_ref, width=200)

        st.markdown("### Sumár podľa formátu")
        st.markdown(style_simple_table(df_fmt_sum, bold_last=True).to_html(), unsafe_allow_html=True)

        # -- SUMÁR podľa turnaja (Rok ↓, Rezort) + Spolu
        rezort_map = {}
        if not df_tournaments.empty and "Rok" in df_tournaments.columns and "Rezort" in df_tournaments.columns:
            rezort_map = {int(r["Rok"]): str(r["Rezort"]).strip() for _, r in df_tournaments.iterrows() if pd.notna(r["Rok"])}

        year_stats = {}
        for _, r in df_player.iterrows():
            y = int(r.get("Rok")) if pd.notna(r.get("Rok")) else None
            if y is None:
                continue
            year_stats.setdefault(y, {"pts": 0.0, "cnt": 0})
            year_stats[y]["pts"] += float(r["_points"])
            year_stats[y]["cnt"] += 1

        rows_years = []
        y_tot_pts, y_tot_cnt = 0.0, 0
        for y in sorted(year_stats.keys(), reverse=True):
            pts = float(year_stats[y]["pts"]); cnt = int(year_stats[y]["cnt"])
            rows_years.append({
                "Rok": y,
                "Rezort": rezort_map.get(y, ""),
                "Body": _fmt_pts(pts),
                "Zápasy": cnt,
                "Úspešnosť": f"{_pct(pts, cnt)} %"
            })
            y_tot_pts += pts; y_tot_cnt += cnt

        rows_years.append({"Rok": "", "Rezort": "Spolu", "Body": _fmt_pts(y_tot_pts), "Zápasy": y_tot_cnt, "Úspešnosť": f"{_pct(y_tot_pts, y_tot_cnt)} %"})
        df_year_sum = pd.DataFrame(rows_years, columns=["Rok", "Rezort", "Body", "Zápasy", "Úspešnosť"])

        st.markdown("### Sumár podľa turnaja")
        df_year_sum_disp = df_year_sum.copy()
        if _device_type == 'mobil':
            df_year_sum_disp = df_year_sum_disp.rename(columns={'Body':'B','Zápasy':'Z','Úspešnosť':'Ú'})
            st.markdown('<div class="mobile-fit">', unsafe_allow_html=True)
        st.markdown(style_simple_table(df_year_sum_disp, bold_last=True).to_html(), unsafe_allow_html=True)
        if _device_type == 'mobil':
            st.markdown('</div>', unsafe_allow_html=True)

        # -- TABUĽKA PÁROV ROZDELENÁ NA 2 STĹPCE: Foursome / Fourball (iba strana vybraného hráča)
            df_pairs = df_player[df_player["Formát"].isin(["Foursome", "Fourball"])].copy() if not df_player.empty else df_player.copy()

            if not df_pairs.empty:
                pair_col = "Lefties" if player_team == "Lefties" else "Righties"

                # Fallback – ak v dátach chýbajú Lefties/Righties, zlož ich z L1/L2 a R1/R2
                if pair_col == "Lefties" and "Lefties" not in df_pairs.columns and {"L1", "L2"}.issubset(df_pairs.columns):
                    df_pairs["Lefties"] = df_pairs[["L1", "L2"]].astype(str).agg(", ".join, axis=1)
                if pair_col == "Righties" and "Righties" not in df_pairs.columns and {"R1", "R2"}.issubset(df_pairs.columns):
                    df_pairs["Righties"] = df_pairs[["R1", "R2"]].astype(str).agg(", ".join, axis=1)

                # Čistič názvov dvojíc
                def _clean_pair_name(x) -> str:
                    if isinstance(x, (list, tuple)):
                        return ", ".join(map(str, x)).strip()
                    s = str(x).strip()
                    if s.startswith("(") and s.endswith(")"):
                        s = s[1:-1].strip()
                    if s.endswith(","):
                        s = s[:-1].strip()
                    s = s.strip("'").strip('"')
                    s = re.sub(r"\s*,\s*", ", ", s)
                    return s

                def _pairs_table_for_format(df_src: pd.DataFrame, fmt_name: str) -> pd.DataFrame:
                    sub = df_src[df_src["Formát"] == fmt_name].copy()
                    if sub.empty:
                        return pd.DataFrame(columns=[pair_col, "Body", "Zápasy", "Úspešnosť"])

                    g = sub.groupby([pair_col], dropna=False)
                    rows = []
                    tot_pts, tot_cnt = 0.0, 0
                    for pnames, grp in g:
                        pts = float(grp["_points"].sum())
                        cnt = int(len(grp))
                        succ = int(round((pts / cnt) * 100)) if cnt else 0
                        rows.append({pair_col: pnames, "_Body_num": pts, "_Zápasy_num": cnt, "_Úspešnosť_num": succ})
                        tot_pts += pts
                        tot_cnt += cnt

                    out = pd.DataFrame(rows)
                    if out.empty:
                        return pd.DataFrame(columns=[pair_col, "Body", "Zápasy", "Úspešnosť"])

                    out[pair_col] = out[pair_col].apply(_clean_pair_name)
                    out.sort_values(by=["_Úspešnosť_num", "_Body_num"], ascending=[False, False], inplace=True)

                    def _fmt_pts(x: float) -> str:
                        return f"{int(x)}" if float(x).is_integer() else f"{x:.1f}"

                    out["Body"] = out["_Body_num"].apply(_fmt_pts)
                    out["Zápasy"] = out["_Zápasy_num"].astype(int)
                    out["Úspešnosť"] = out["_Úspešnosť_num"].astype(int).map(lambda v: f"{v} %")
                    out = out[[pair_col, "Body", "Zápasy", "Úspešnosť"]]

                    succ_tot = int(round((tot_pts / tot_cnt) * 100)) if tot_cnt else 0
                    out = pd.concat([
                        out,
                        pd.DataFrame([{pair_col: "Spolu", "Body": _fmt_pts(tot_pts), "Zápasy": tot_cnt, "Úspešnosť": f"{succ_tot} %"}])
                    ], ignore_index=True)
                    return out

                df_pairs_fs = _pairs_table_for_format(df_pairs, "Foursome")
                df_pairs_fb = _pairs_table_for_format(df_pairs, "Fourball")

                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("### Dvojice Foursome")
                    if df_pairs_fs.empty:
                        st.info("Žiadne párové zápasy vo formáte **Foursome**.")
                    else:
                        st.markdown(style_simple_table(df_pairs_fs, bold_last=True).to_html(), unsafe_allow_html=True)

                with c2:
                    st.markdown("### Dvojice Fourball")
                    if df_pairs_fb.empty:
                        st.info("Žiadne párové zápasy vo formáte **Fourball**.")
                    else:
                        st.markdown(style_simple_table(df_pairs_fb, bold_last=True).to_html(), unsafe_allow_html=True)
            else:
                st.info("Hráč neodohral žiadne zápasy formátov **Foursome/Fourball** v zvolených rokoch.")

            # --- PROTIHRÁČI# --- PROTIHRÁČI (agregácia za všetkých súperov z aktuálne vyfiltrovaných zápasov hráča)
            # Ak je hráč v Lefties -> súperi z R1/R2, ak v Righties -> súperi z L1/L2
            opp_cols = ["R1", "R2"] if player_team == "Lefties" else ["L1", "L2"]

            if not df_player.empty:
                # nazbieraj W/D/L + body a počty zápasov proti jednotlivým súperom
                # kľúč = celé meno protihráča v kanonickej podobe z dát (Priezvisko Meno)
                agg = {}  # {opp_name: {"wins":int, "draws":int, "losses":int, "pts":float, "cnt":int}}

                for _, row in df_player.iterrows():
                    # zistime stranu vybraného hráča v danom zápase
                    is_left = (str(row.get("L1","")).strip() == selected_canonical) or (str(row.get("L2","")).strip() == selected_canonical)
                    is_right = (str(row.get("R1","")).strip() == selected_canonical) or (str(row.get("R2","")).strip() == selected_canonical)

                    # body strán v danom zápase (na istotu ako float)
                    lb = float(row.get("Lbody", 0) or 0.0)
                    rb = float(row.get("Rbody", 0) or 0.0)

                    # hráčove body a súperove body (z pohľadu výsledku)
                    if is_left and not is_right:
                        my_pts = lb
                        opp_pts = rb
                        opponents = [row.get("R1", None), row.get("R2", None)]
                    elif is_right and not is_left:
                        my_pts = rb
                        opp_pts = lb
                        opponents = [row.get("L1", None), row.get("L2", None)]
                    else:
                        # fallback (nemalo by sa stať): využi vypočítané _points a opačnú stranu podľa player_team
                        my_pts = float(row.get("_points", 0.0) or 0.0)
                        opponents = [row.get(c, None) for c in (opp_cols)]
                        opp_pts = rb if player_team == "Lefties" else lb

                    # urč výsledok W/D/L pre tento riadok
                    if my_pts > opp_pts:
                        res = "win"
                    elif my_pts < opp_pts:
                        res = "loss"
                    else:
                        res = "draw"

                    # zozbieraj mená súperov (ignoruj NaN/prázdne), každý dostane rovnaký výsledok
                    for nm in opponents:
                        if pd.isna(nm) or str(nm).strip() == "":
                            continue
                        opp_name = str(nm).strip()
                        bucket = agg.setdefault(opp_name, {"wins": 0, "draws": 0, "losses": 0, "pts": 0.0, "cnt": 0})
                        # výsledok
                        if res == "win":
                            bucket["wins"] += 1
                        elif res == "loss":
                            bucket["losses"] += 1
                        else:
                            bucket["draws"] += 1
                        # body + zápasy z pohľadu vybraného hráča
                        bucket["pts"] += my_pts
                        bucket["cnt"] += 1

                # zostav DataFrame
                rows = []
                tot_pts, tot_cnt = 0.0, 0
                tot_w, tot_d, tot_l = 0, 0, 0
                for opp, d in agg.items():
                    pts = float(d["pts"])
                    cnt = int(d["cnt"])
                    w, d_, l = int(d["wins"]), int(d["draws"]), int(d["losses"])
                    succ = int(round((pts / cnt) * 100)) if cnt else 0
                    rows.append({
                        "Protihráč": to_firstname_first(opp),   # celé meno (Meno Priezvisko)
                        "Výhra": w,
                        "Remíza": d_,
                        "Prehra": l,
                        "_Body_num": pts,
                        "_Zápasy_num": cnt,
                        "_Úspešnosť_num": succ,
                    })
                    tot_pts += pts
                    tot_cnt += cnt
                    tot_w += w
                    tot_d += d_
                    tot_l += l

                import pandas as pd
                df_opp = pd.DataFrame(rows)

                # Nadpis tabuľky zobraziť priamo s menom hráča
                st.markdown(f"### {selected_display} a protihráči")

                if df_opp.empty:
                    st.info("V zvolených zápasoch sa nenašli žiadni protihráči.")
                else:
                    # zoradenie: Úspešnosť ↓, Body ↓, Protihráč ↑
                    df_opp.sort_values(
                        by=["_Úspešnosť_num", "_Body_num", "Protihráč"],
                        ascending=[False, False, True],
                        inplace=True
                    )

                    # formátovanie výstupu
                    def _fmt_pts(x: float) -> str:
                        return f"{int(x)}" if float(x).is_integer() else f"{x:.1f}"

                    df_opp["Body"] = df_opp["_Body_num"].apply(_fmt_pts)
                    df_opp["Zápasy"] = df_opp["_Zápasy_num"].astype(int)
                    df_opp["Úspešnosť"] = df_opp["_Úspešnosť_num"].astype(int).map(lambda v: f"{v} %")

                    # finálny výber stĺpcov – bez „Spolu“
                    df_opp_disp = df_opp[["Protihráč", "Výhra", "Remíza", "Prehra", "Body", "Zápasy", "Úspešnosť"]].copy()
                    # Pre export: vždy desktop reprezentácia (aj na mobile)
                    df_opp_export = df_opp_disp.copy()
                    if _device_type == 'mobil':
                        df_opp_disp['Protihráč'] = df_opp_disp['Protihráč'].apply(short_name_msurname)
                        df_opp_disp['V-A/S-P'] = df_opp_disp.apply(lambda rr: str(int(rr['Výhra']))+'-'+str(int(rr['Remíza']))+'-'+str(int(rr['Prehra'])), axis=1)
                        df_opp_disp = df_opp_disp.rename(columns={'Body':'B','Zápasy':'Z','Úspešnosť':'Ú'})
                        df_opp_disp = df_opp_disp[["Protihráč", "V-A/S-P", "B", "Z", "Ú"]]
                        st.markdown('<div class="mobile-fit">', unsafe_allow_html=True)
                    st.markdown(style_simple_table(df_opp_disp, bold_last=False).to_html(), unsafe_allow_html=True)
                    if _device_type == 'mobil':
                        st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.info("Hráč nemá v zvolených **rokoch** a vybraných **formátoch** žiadne zápasy (pre výpočet protihráčov).")
        

            
            # -- ZÁPASY: zoradenie Rok ↓, Deň ↑, Zápas ↑ a render
            if not df_player.empty:
                if "Rok" in df_player.columns:
                    df_player["Rok"] = pd.to_numeric(df_player["Rok"], errors="coerce").astype("Int64")
                if "Deň" in df_player.columns:
                    day_clean = df_player["Deň"].astype(str).str.strip().str.replace(r"\.$", "", regex=True)
                    df_player["_day_int"] = pd.to_numeric(day_clean, errors="coerce").fillna(0).astype(int)
                else:
                    df_player["_day_int"] = 0

                by = []; ascending = []
                if "Rok" in df_player.columns: by.append("Rok"); ascending.append(False)   # desc
                by.append("_day_int"); ascending.append(True)
                if "Zápas" in df_player.columns: by.append("Zápas"); ascending.append(True)

                df_player.sort_values(by=by, ascending=ascending, inplace=True)
                df_player.drop(columns=["_day_int"], inplace=True, errors="ignore")

                wanted_cols = ["Rok", "Deň", "Zápas", "Formát", "Lefties", "Righties", "Víťaz"]
                cols_present = [c for c in wanted_cols if c in df_player.columns]
                matches_view = df_player[cols_present].copy()

                if _device_type == 'mobil':
                    mv = matches_view.copy()
                    fmt_map = {'Foursome':'Fs','Fourball':'Fb','Single':'S'}

                    def _int_str(v):
                        try:
                            return str(int(float(v)))
                        except Exception:
                            s = str(v).strip()
                            return s.replace('.', '') if s.endswith('.') else s

                    f_abbr = mv['Formát'].astype(str).map(lambda x: fmt_map.get(x, x)) if 'Formát' in mv.columns else ''
                    mv['Zápas'] = (
                        mv['Rok'].map(_int_str) + '-' +
                        mv['Deň'].map(_int_str) + '-' +
                        mv['Zápas'].map(_int_str) + '-' +
                        f_abbr
                    )

                    # ❗Lefties/Righties nechaj rovnaké ako desktop (bez skratiek mien)
                    # if 'Lefties' in mv.columns: mv['Lefties'] = mv['Lefties'].apply(short_pair_names)
                    # if 'Righties' in mv.columns: mv['Righties'] = mv['Righties'].apply(short_pair_names)

                    # Víťaz skráť na V a hodnoty Lefties/Righties na L/R ponechaj (toto je OK)
                    if 'Víťaz' in mv.columns:
                        mv['Víťaz'] = mv['Víťaz'].astype(str).str.replace('Lefties','L').str.replace('Righties','R')
                        mv.rename(columns={'Víťaz': 'V'}, inplace=True)

                    # ✅ Stĺpce ponechaj: Zápas, Lefties, Righties, V, A/S
                    cols = ['Zápas'] + [c for c in ['Lefties','Righties','V','A/S'] if c in mv.columns]
                    matches_view = mv[cols].copy()
                
                sty = style_matches_table(matches_view)
                st.markdown("### Zápasy")
                if _device_type == 'mobil':
                    st.markdown('<div class="mobile-fit">', unsafe_allow_html=True)
                st.markdown(sty.to_html(), unsafe_allow_html=True)
                if _device_type == 'mobil':
                    st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.info("Hráč nemá v zvolených **rokoch** a vybraných **formátoch** žiadne zápasy.")

            # --- Export DETAIL HRÁČA do Excelu (1 hárok na každú tabuľku + Filter) ---
            try:
                # 0) Helper: zápis DF do hárka + centrovanie + autofit
                def _write_sheet_auto_fit(writer, df: pd.DataFrame, sheet_name: str, default_cols: list[str] | None = None):
                    # ak DF chýba / je prázdny, založ prázdny s očakávanými hlavičkami (nech je štruktúra stabilná)
                    if df is None or df.empty:
                        df_to_save = pd.DataFrame(columns=default_cols or [])
                    else:
                        df_to_save = df.copy()

                    df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws = writer.sheets[sheet_name]

                    # centrovanie všetkých buniek (hlavička + dáta)
                    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=r, column=c).alignment = align_center

                    # autofit šírky: max(dĺžka hlavičky, dĺžka obsahu) + padding
                    for col_idx, col_name in enumerate(df_to_save.columns, start=1):
                        series = df_to_save[col_name].astype(str).fillna("")
                        max_len = max([len(str(col_name))] + series.map(len).tolist()) if not series.empty else len(str(col_name))
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)  # bezpečnostný limit

                # 1) Priprav mapu -> DF pre všetky tabuľky v Detaily hráča
                # Pozn.: Niektoré premenne vznikajú len ak existujú dáta – preto používame locals().get(...)
                pair_col_name = "Lefties" if player_team == "Lefties" else "Righties"

                sheets: dict[str, pd.DataFrame] = {
                    "Sumár podľa formátu": locals().get("df_fmt_sum", pd.DataFrame()),
                    "Sumár podľa turnaja": locals().get("df_year_sum", pd.DataFrame()),
                    "Dvojice Foursome":     locals().get("df_pairs_fs", pd.DataFrame()),
                    "Dvojice Fourball":     locals().get("df_pairs_fb", pd.DataFrame()),
                    "Protihráči": locals().get("df_opp_export", locals().get("df_opp_disp", pd.DataFrame())),
                    "Zápasy": locals().get("matches_view_export", locals().get("matches_view", pd.DataFrame())),
                }

                # 2) Hárok FILTER – vypíš všetky vybraté položky
                rows_filter = []
                if getattr(FILTER, "t_all", False):
                    rows_filter.append({"Kategória": "Turnaje", "Hodnota": "všetky turnaje"})
                else:
                    for lbl in getattr(FILTER, "t_selected", []):
                        rows_filter.append({"Kategória": "Turnaj", "Hodnota": str(lbl)})

                teams_val = ", ".join(getattr(FILTER, "teams", [])) if getattr(FILTER, "teams", []) else "—"
                fmts_val  = ", ".join(getattr(FILTER, "formats", [])) if getattr(FILTER, "formats", []) else "—"
                rows_filter.append({"Kategória": "Tímy", "Hodnota": teams_val})
                rows_filter.append({"Kategória": "Formáty", "Hodnota": fmts_val})
                df_filter_export = pd.DataFrame(rows_filter, columns=["Kategória", "Hodnota"])

                # 3) Pomocný slovník defaultných hlavičiek (ak by bol niektorý DF prázdny)
                defaults = {
                    "Sumár podľa formátu": ["Formát", "Body", "Zápasy", "Úspešnosť"],
                    "Sumár podľa turnaja": ["Rok", "Rezort", "Body", "Zápasy", "Úspešnosť"],
                    "Dvojice Foursome":    [pair_col_name, "Body", "Zápasy", "Úspešnosť"],
                    "Dvojice Fourball":    [pair_col_name, "Body", "Zápasy", "Úspešnosť"],
                    "Protihráči":          ["Protihráč", "Výhra", "Remíza", "Prehra", "Body", "Zápasy", "Úspešnosť"],
                    "Zápasy":              ["Rok", "Deň", "Zápas", "Formát", "Lefties", "Righties", "Víťaz"],
                    "Filter":              ["Kategória", "Hodnota"],
                }

                # 4) Názov súboru: LR - {Meno Priezvisko} - YYYY.MM.DD-hh.mm.ss.xlsx (očistené od nepovolených znakov)
                timestamp = datetime.now().strftime("%Y.%m.%d-%H.%M.%S")
                safe_player = re.sub(r'[\\/:*?"<>|]+', " ", selected_display).strip()
                xlsx_name = f"LR - {safe_player} - {timestamp}.xlsx"

                # 5) Export do pamäte a download tlačidlo
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    # najprv všetky tabuľky
                    for sheet_name, df_ in sheets.items():
                        _write_sheet_auto_fit(writer, df_, sheet_name, default_cols=defaults.get(sheet_name))
                    # nakoniec FILTER
                    _write_sheet_auto_fit(writer, df_filter_export, "Filter", default_cols=defaults["Filter"])

                st.download_button(
                    label=f"⬇️ Export detailu hráča do Excelu ({xlsx_name})",
                    data=buffer.getvalue(),
                    file_name=xlsx_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"dl_xlsx_player_{safe_player}",
                )

            except Exception as _ex:
                st.warning(f"Export detailu hráča do Excelu sa nepodaril: {type(_ex).__name__}: {_ex}")
    




# *****************************
# Turnaje
# *****************************
with tab_turnaje:
    st.subheader("Turnaje")
    tdf = df_tournaments.copy()
    if "Rok" in tdf.columns:
        tdf.sort_values("Rok", ascending=False, inplace=True)

    if 'open_year' not in st.session_state:
        st.session_state['open_year'] = None

    for _, t in tdf.iterrows():
        year = int(t.get('Rok')) if pd.notna(t.get('Rok')) else None
        rezort = str(t.get('Rezort', '')).strip()
        l_captain = str(t.get('L-Captain', '')).strip()
        r_captain = str(t.get('R-Captain', '')).strip()
        winner_val = str(t.get('Víťaz', '')).strip().lower()
        btn_icon = '🔵' if winner_val == 'lefties' else ('🔴' if winner_val == 'righties' else '⚪')
        btn_label = f"{btn_icon}    {year}     {rezort}"
        clicked = st.button(btn_label, key=f"btn_{year}")
        if clicked:
            st.session_state['open_year'] = year if st.session_state.get('open_year') != year else None
        if st.session_state.get('open_year') == year:
            logo_url = str(t.get('Logo', '')).strip()
            if logo_url:
                st.image(logo_url, width=240)

            df_y = df_matches[df_matches['Rok'] == year].copy()
            l_total = float(df_y['Lbody'].fillna(0).sum()) if 'Lbody' in df_y.columns else 0.0
            r_total = float(df_y['Rbody'].fillna(0).sum()) if 'Rbody' in df_y.columns else 0.0

            def _fmt(v: float) -> str:
                return f"{int(v)}" if float(v).is_integer() else f"{v:.1f}"

            st.markdown(f"**Výsledok turnaja {year}:** Lefties **{_fmt(l_total)}** : **{_fmt(r_total)}** Righties")

            val_L = t.get('StavL', t.get('Stav L', None))
            val_R = t.get('StavR', t.get('Stav R', None))
            try:
                val_L = float(val_L) if val_L is not None else 0.0
            except Exception:
                val_L = 0.0
            try:
                val_R = float(val_R) if val_R is not None else 0.0
            except Exception:
                val_R = 0.0
            st.markdown(f"**Stav na konci turnaja {year}:** Lefties **{_fmt(val_L)}** : **{_fmt(val_R)}** Righties")

            df_y = df_matches[df_matches['Rok'] == year].copy()
            left_players, right_players = players_for_year_pairs_only(df_y)
            left_table = build_team_table(df_y, left_players, side='L')
            right_table = build_team_table(df_y, right_players, side='R')
            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"### Team Lefties {year}  \n(kapitán: {to_firstname_first(l_captain)})")
                if not left_table.empty:
                    if _device_type == "mobil" and "Hráč" in left_table.columns:
                        left_table = left_table.copy()
                        left_table["Hráč"] = left_table["Hráč"].apply(short_name_msurname)
                    sty = style_team_table(left_table, "L")
                    st.markdown(f"{sty.to_html()}", unsafe_allow_html=True)
                else:
                    st.info("Pre tento rok nie sú v dátach hráči tímu Lefties.")
            with c2:
                st.markdown(f"### Team Righties {year}  \n(kapitán: {to_firstname_first(r_captain)})")
                if not right_table.empty:
                    if _device_type == "mobil" and "Hráč" in right_table.columns:
                        right_table = right_table.copy()
                        right_table["Hráč"] = right_table["Hráč"].apply(short_name_msurname)
                    sty = style_team_table(right_table, "R")
                    st.markdown(f"{sty.to_html()}", unsafe_allow_html=True)
                else:
                    st.info("Pre tento rok nie sú v dátach hráči tímu Righties.")

            st.markdown("---")
            wanted_cols = ["Rok", "Deň", "Zápas", "Formát", "Lefties", "Righties", "Víťaz", "A/S"]
            cols_present = [c for c in wanted_cols if c in df_y.columns]
            matches_view = df_y[cols_present].copy()
            # Pre export: vždy desktop reprezentácia (aj na mobile)
            matches_view_export = matches_view.copy()
            # Turnaje: vždy všetky zápasy za vybraný rok (ignoruje Filter aj Detail hráča)

            if _device_type == 'mobil':
                mv = matches_view.copy()
                fmt_map = {'Foursome':'Fs','Fourball':'Fb','Single':'S'}

                def _int_str(v):
                    try:
                        return str(int(float(v)))
                    except Exception:
                        s = str(v).strip()
                        return s.replace('.', '') if s.endswith('.') else s

                f_abbr = mv['Formát'].astype(str).map(lambda x: fmt_map.get(x, x)) if 'Formát' in mv.columns else ''
                mv['Zápas'] = mv['Deň'].map(_int_str) + '-' + mv['Zápas'].map(_int_str) + '-' + f_abbr

                # ❗Lefties/Righties nechaj ako desktop (bez skratiek mien)
                # if 'Lefties' in mv.columns: mv['Lefties'] = mv['Lefties'].apply(short_pair_names)
                # if 'Righties' in mv.columns: mv['Righties'] = mv['Righties'].apply(short_pair_names)

                # Víťaz skráť na V a hodnoty Lefties/Righties na L/R ponechaj
                if 'Víťaz' in mv.columns:
                    mv['Víťaz'] = mv['Víťaz'].astype(str).str.replace('Lefties','L').str.replace('Righties','R')
                    mv.rename(columns={'Víťaz': 'V'}, inplace=True)

                # ✅ Stĺpce ponechaj: Zápas, Lefties, Righties, V, A/S
                cols = ['Zápas'] + [c for c in ['Lefties','Righties','V','A/S'] if c in mv.columns]
                matches_view = mv[cols].copy()

                st.markdown(f"### Zápasy {year}")
            sty = style_matches_table(matches_view)
            if _device_type == 'mobil':
                st.markdown('<div class="mobile-fit">', unsafe_allow_html=True)
            st.markdown(sty.to_html(), unsafe_allow_html=True)
            if _device_type == 'mobil':
                st.markdown('</div>', unsafe_allow_html=True)

            # --- Export do Excelu: Team Lefties {year}, Team Righties {year}, Zápasy {year} ---
            try:
                # Priprav názov súboru: L&R {Rok} {Rezort}.xlsx (bez neplatných znakov)
                safe_rezort = re.sub(r'[\\/:*?"<>|]+', ' ', rezort).strip()
                timestamp = datetime.now().strftime("%Y.%m.%d-%H.%M.%S")
                xlsx_name = f"L&R - {year} - {safe_rezort} ({timestamp}).xlsx"

                # Funkcia na export DF -> hárok + autofit stĺpcov
                def _write_sheet_auto_fit(writer, df: pd.DataFrame, sheet_name: str):
                    # Ak je DF prázdny, exportuj aspoň hlavičky (nech má hárok konzistentnú štruktúru)
                    if df is None or df.empty:
                        if sheet_name.startswith("Team "):
                            df_export = pd.DataFrame(columns=["Hráč", "Body", "Zápasy", "Úspešnosť"])
                        else:
                            df_export = pd.DataFrame(columns=["Rok", "Deň", "Zápas", "Formát", "Lefties", "Righties", "Víťaz"])
                    else:
                        df_export = df.copy()

                    # Zapíš dáta
                    df_export.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Vezmi worksheet a centrovanie + autofit
                    ws = writer.sheets[sheet_name]

                    # 1) Centrovanie všetkých buniek (hlavičky aj obsah)
                    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
                    max_row = ws.max_row
                    max_col = ws.max_column
                    for r in range(1, max_row + 1):
                        for c in range(1, max_col + 1):
                            ws.cell(row=r, column=c).alignment = align_center

                    # 2) Auto-fit šírky stĺpcov podľa najdlhšieho textu v stĺpci (vrátane hlavičky)
                    for col_idx, col_name in enumerate(df_export.columns, start=1):
                        series = df_export[col_name].astype(str).fillna("")
                        max_len = max([len(str(col_name))] + series.map(len).tolist())
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)  # bezpečnostný limit
        
                # Zostav DF pre export
                sheet_left  = left_table.copy()  if 'left_table'  in locals() else pd.DataFrame()
                sheet_right = right_table.copy() if 'right_table' in locals() else pd.DataFrame()
                sheet_games = matches_view_export.copy() if 'matches_view_export' in locals() else (matches_view.copy() if 'matches_view' in locals() else pd.DataFrame())

                # (Voliteľné) zoradenie stĺpcov, ak by DF prišli v inom poradí
                # Team hárky: Hráč, Body, Zápasy, Úspešnosť
                for _df in (sheet_left, sheet_right):
                    if not _df.empty:
                        cols = [c for c in ["Hráč","Body","Zápasy","Úspešnosť"] if c in _df.columns]
                        if cols:
                            _df = _df[cols]
                # Zápasy: Rok, Deň, Zápas, Formát, Lefties, Righties, Víťaz
                if not sheet_games.empty:
                    cols = [c for c in ["Rok","Deň","Zápas","Formát","Lefties","Righties","Víťaz"] if c in sheet_games.columns]
                    if cols:
                        sheet_games = sheet_games[cols]

                # Export do pamäte
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    _write_sheet_auto_fit(writer, sheet_left,  f"Team Lefties {year}")
                    _write_sheet_auto_fit(writer, sheet_right, f"Team Righties {year}")
                    _write_sheet_auto_fit(writer, sheet_games, f"Zápasy {year}")

                st.download_button(
                    label=f"⬇️ Export do Excelu ({xlsx_name})",
                    data=buffer.getvalue(),
                    file_name=xlsx_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"dl_xlsx_{year}",
                )
            except Exception as _ex:
                st.warning(f"Export do Excelu sa nepodaril: {type(_ex).__name__}: {_ex}")
                
            
            photo_url = str(t.get('Photo', '')).strip()
            if photo_url:
                st.image(photo_url,  width=800)
            #     st.image(photo_url,  use_container_width=True)
            st.markdown("")




# *****************************
# Filter
# *****************************
with tab_filter:
    st.subheader("Filter")
    c1, c2 = st.columns([2, 1])

    with c1:
        st.markdown("### Turnaje")
        tournament_items = _build_tournament_items(df_tournaments)
        st.session_state.setdefault('flt_t_keys', [it['key'] for it in tournament_items])

        # Master
        st.checkbox("Všetky turnaje", key='flt_t_all', on_change=_toggle_all_tournaments)

        # Deti
        for item in tournament_items:
            st.session_state.setdefault(item['key'], True)
            st.checkbox(item['label'], key=item['key'], on_change=_on_filter_change)

        selected_tournaments = [it['label'] for it in tournament_items if st.session_state.get(it['key'], False)]
        st.session_state['flt_tournaments'] = selected_tournaments
        st.caption(f"Vybrané turnaje: {len(selected_tournaments)}/{len(tournament_items)}")

    with c2:
        st.markdown("### Tímy")
        st.checkbox("Lefties", key='flt_team_lefties', on_change=_on_filter_change)
        st.checkbox("Righties", key='flt_team_righties', on_change=_on_filter_change)

        st.markdown("### Formáty hry")
        st.checkbox("Foursome", key='flt_fmt_foursome', on_change=_on_filter_change)
        st.checkbox("Fourball", key='flt_fmt_fourball', on_change=_on_filter_change)
        st.checkbox("Single", key='flt_fmt_single', on_change=_on_filter_change)
